"""For reference only, do not use

This script dumped all the codelists in the previous Excel template to SKOS RDF files."""

import datetime
from rdflib import Graph
from openpyxl import load_workbook as _load_workbook
from pathlib import Path

wb = _load_workbook(filename="../examples/data-submission-template-minerals-3.0-001.xlsx", data_only=True)

VOCAB_DIR = Path(__file__).parent / "vocabs-codelists"

ws = wb["VALIDATION_DICTIONARY"]

prefixes = """PREFIX : <https://linked.data.gov.au/def/gsq-geochem/{id}/>
PREFIX agldwgstatus: <https://linked.data.gov.au/def/reg-statuses/>
PREFIX cs: <https://linked.data.gov.au/def/gsq-geochem/{id}>
PREFIX dcterms: <http://purl.org/dc/terms/>
PREFIX owl: <http://www.w3.org/2002/07/owl#>
PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
PREFIX reg: <http://purl.org/linked-data/registry#>
PREFIX sampty: <https://linked.data.gov.au/def/sample-type/>
PREFIX sdo: <https://schema.org/>
PREFIX skos: <http://www.w3.org/2004/02/skos/core#>
PREFIX xsd: <http://www.w3.org/2001/XMLSchema#>

"""

concept_scheme = """

cs: 
    a skos:ConceptScheme ;
    skos:notation "{list}" ;
    skos:prefLabel "{label}"@en ;
    skos:definition "{label}"@en ;
    sdo:creator <https://linked.dat.gov.au/org/gsq> ;
    sdo:publisher <https://linked.dat.gov.au/org/gsq> ;
    sdo:dateCreated "2023-11-11"^^xsd:date ;
    sdo:dateModified "{todays_date}"^^xsd:date ;
    skos:historyNote "Created from the Geochemistry data exchange Excel Template v2, list {list} November 2023" ;
    sdo:copyrightNotice "(c) Geological Survey of Queensland, 2023" ;
    sdo:license <https://purl.org/NET/rdflicense/cc-by4.0> ;
    reg:status agldwgstatus:experimental ;
.

<https://linked.dat.gov.au/org/gsq>
    a sdo:Organization ;
    sdo:name "Geological Survey of Queensland" ;
    sdo:url "https://www.business.qld.gov.au/industries/mining-energy-water/resources/geoscience-information/gsq"^^xsd:anyURI ;
.
"""

concept_template = """
:{iri}
    a skos:Concept ;
    rdfs:isDefinedBy cs: ;
    skos:definition 'Definition for {label}'@en ;
    skos:inScheme cs: ;
    skos:notation '{notation}' ;
    skos:prefLabel '{label}'@en ;
    skos:topConceptOf cs: ;
    skos:historyNote "Taken from Geochem Excel exchange template v2" ;
.

cs: skos:hasTopConcept :{iri} .
"""


def id_to_iri(id: str):
    replacements = {
        ' ': '-',
        '(': '',
        ')': '',
        '/': '-',
        '±': '',
        ',': '',
        '#': 'No',
        '"': "in",
        '.': ''
    }
    replaced_chars = [replacements.get(char, char) for char in id.lower()]
    return ''.join(replaced_chars).replace("--", "-")


def id_to_label(id: str):
    id = id.strip().lower()
    replacements = {
        'xxx': "in",
    }
    replaced_chars = [replacements.get(char, char) for char in id.lower()]
    return ''.join(replaced_chars).replace("--", "-")


col = 0
while True:
    col += 1
    val = ws.cell(row=4, column=col).value

    if val is None:
        break
    else:
        print(f"Column {val}")
        voc_file_name = str(val).lower().replace(" ", "-").replace("_", "-")
        vocab_label = str(val).title().replace("_", " ")
        print(voc_file_name)
        grf = ""
        grf += prefixes.replace("{id}", voc_file_name)
        grf += concept_scheme\
            .replace("{label}", vocab_label)\
            .replace("{todays_date}", datetime.datetime.now().strftime("%Y-%m-%d"))\
            .replace("{list}", val)

        row = 5
        while True:
            val = ws.cell(row=row, column=col).value
            if val is None:
                break
            else:
                row += 1
                label = id_to_label(val)
                iri = id_to_iri(label)
                grf += concept_template\
                    .replace("{iri}", iri)\
                    .replace("{label}", label)\
                    .replace("{notation}", val)

        Graph().parse(data=grf, format="turtle").serialize(destination=VOCAB_DIR / f"{voc_file_name}.ttl", format="longturtle")
