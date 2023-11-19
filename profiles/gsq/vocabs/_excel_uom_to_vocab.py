"""For reference only, do not use

This script dumped the UoMs from the UNITS_OF_MEASURE Ecel worksheet into a simplified CSV file for RDF creation."""

import csv
import datetime
from pathlib import Path

from openpyxl import load_workbook as _load_workbook
from rdflib import Graph, Namespace, URIRef, Literal
from rdflib.namespace import RDF, RDFS, SKOS

UOM = Namespace("https://linked.data.gov.au/def/gsq-geochem/uom/")

VOCAB_DIR = Path(__file__).parent / "vocabs-uom"
OUTPUT_FILE = VOCAB_DIR / "uom.ttl"

wb = _load_workbook(filename="../examples/data-submission-template-minerals-3.0-001.xlsx", data_only=True)
ws = wb["UNITS_OF_MEASURE"]


def id_to_iri(id: str):
    replacements = {
        ' ': '-',
        '(': '',
        ')': '',
        '/': '-',
        '±': '',
        ',': '',
        '#': 'No',
        '"': "in",
        '.': ''
    }
    replaced_chars = [replacements.get(char, char) for char in id.lower()]
    return ''.join(replaced_chars).replace("--", "-")

concepts = set()

col = 0
while True:
    col += 1
    val = ws.cell(row=1, column=col).value

    if val is None:
        break
    else:
        print(f"Column {val}")
        collection_name = str(val).title().replace("_", " ")
        collection_iri = id_to_iri(val)

        row = 1
        while True:
            val = ws.cell(row=row, column=col).value
            if val is None:
                break
            else:
                row += 1
                val = ws.cell(row=row, column=col).value
                if val is None:
                    break
                concept_code = str(val).split("(")[1].split(")")[0]
                concepts.add((concept_code, collection_iri))

with open("vocabs-uom/uom.csv", "w") as f:
    f.write(f"ExcelCode,Collection\n")
    for concept in sorted(concepts):
        f.write(f"{concept[0]},{concept[1]}\n")


UOM = Namespace("https://linked.data.gov.au/def/gsq-geochem/uom/")

VOCAB_DIR = Path(__file__).parent / "vocabs-uom"
OUTPUT_FILE = VOCAB_DIR / "uom-new.ttl"

header = """PREFIX : <https://linked.data.gov.au/def/gsq-geochem/uom/>
PREFIX agldwgstatus: <https://linked.data.gov.au/def/reg-statuses/>
PREFIX cs: <https://linked.data.gov.au/def/gsq-geochem/uom>
PREFIX dcterms: <http://purl.org/dc/terms/>
PREFIX owl: <http://www.w3.org/2002/07/owl#>
PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
PREFIX reg: <http://purl.org/linked-data/registry#>
PREFIX sdo: <https://schema.org/>
PREFIX skos: <http://www.w3.org/2004/02/skos/core#>
PREFIX unit: <http://qudt.org/vocab/unit/>
PREFIX units: <http://qudt.org/vocab/unit>
PREFIX xsd: <http://www.w3.org/2001/XMLSchema#>

cs: 
    a skos:ConceptScheme ;
    skos:prefLabel "Geochemistry Units of Measure"@en ;
    skos:definition "The Units of Measure as used in Geochemistry data exchange with GSQ"@en ;
    sdo:creator <https://linked.dat.gov.au/org/gsq> ;
    sdo:publisher <https://linked.dat.gov.au/org/gsq> ;
    sdo:dateCreated "2023-11-11"^^xsd:date ;
    sdo:dateModified "{todays_date}"^^xsd:date ;
    skos:historyNote "Created from the Geochemistry data exchange Excel Template v2, UNITS_OF_MEASURE sheet November 2023" ;
    sdo:copyrightNotice "(c) Geological Survey of Queensland, 2023" ;
    sdo:license <https://purl.org/NET/rdflicense/cc-by4.0> ;
    reg:status agldwgstatus:experimental ;
.

<https://linked.dat.gov.au/org/gsq>
    a sdo:Organization ;
    sdo:name "Geological Survey of Queensland" ;
    sdo:url "https://www.business.qld.gov.au/industries/mining-energy-water/resources/geoscience-information/gsq"^^xsd:anyURI ;
.
"""

g = Graph()
g.parse(data=header.replace("{todays_date}", datetime.datetime.now().strftime("%Y-%m-%d")), format="turtle")

cs_iri = URIRef("https://linked.data.gov.au/def/gsq-geochem/uom")

with open("vocabs-uom/uom.csv") as f:
    for i, row in enumerate(csv.reader(f)):
        if i == 0:
            continue
        else:
            concept_iri = URIRef(row[1].replace("unit:", "http://qudt.org/vocab/unit/"))
            notation = Literal(row[0])
            g.add((concept_iri, RDF.type, SKOS.Concept))
            g.add((concept_iri, SKOS.topConceptOf, cs_iri))
            g.add((cs_iri, SKOS.hasTopConcept, concept_iri))
            g.add((concept_iri, SKOS.notation, notation))
            g.add((concept_iri, RDFS.isDefinedBy, URIRef("http://qudt.org/vocab/unit")))
            collection_iri = URIRef(UOM[row[2]])
            collection_label = str(row[2]).replace("-", " ").title()
            collection_desc = f"Definition for {collection_label}"
            g.add((collection_iri, RDF.type, SKOS.Collection))
            g.add((collection_iri, SKOS.inScheme, cs_iri))
            g.add((collection_iri, SKOS.prefLabel, Literal(collection_label, lang="en")))
            g.add((collection_iri, SKOS.definition, Literal(collection_desc, lang="en")))
            g.add((collection_iri, SKOS.historyNote, Literal("Collection derived from original spreadsheet columns")))
            g.add((collection_iri, SKOS.member, concept_iri))

g.serialize(destination=OUTPUT_FILE, format="longturtle")
