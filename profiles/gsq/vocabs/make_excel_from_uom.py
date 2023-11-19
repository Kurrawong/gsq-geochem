import csv
from pathlib import Path
from rdflib import Graph, URIRef, Namespace, Literal
from rdflib.namespace import RDF, RDFS, SKOS, SDO, XSD

from openpyxl import Workbook
from openpyxl.styles import Font, Border, PatternFill, Side

OUTPUT_FILE = Path(__file__).parent / "uom-dictionary.xlsx"

# styling
ft = Font(bold=True)
light_yellow = PatternFill("solid", fgColor="00FFFF99")
border_thin = Side(border_style="thin", color="000000")
border_double = Side(border_style="double", color="000000")

wb = Workbook()
ws = wb.active
ws.title = "UNITS_OF_MEASURE"


g = Graph()
g.parse("vocabs-uom/uom.ttl")

col = 1
for coll in g.subjects(RDF.type, SKOS.Collection):
    coll_label = g.value(subject=coll, predicate=SKOS.prefLabel)
    cell = ws.cell(column=col, row=1)
    cell.value = coll_label
    cell.font = ft
    cell.fill = light_yellow
    cell.border = Border(bottom=border_double)

    row = 2
    concepts = []
    for con in g.objects(coll, SKOS.member):
        con_label = g.value(subject=con, predicate=SKOS.prefLabel)
        con_notation = g.value(subject=con, predicate=SKOS.notation)
        concepts.append(f"{con_label} ({con_notation})")
        row += 1

    for i, concept in enumerate(sorted(concepts)):
        ws.cell(column=col, row=i+2).value = concept

    col += 1

wb.save(OUTPUT_FILE)
