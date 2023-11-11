"""For reference only, do not use

This script dumped the UoMs from the UNITS_OF_MEASURE Ecel worksheet into a simplified CSV file for RDF creation."""

from pathlib import Path

from openpyxl import load_workbook as _load_workbook
from rdflib import Namespace

UOM = Namespace("https://linked.data.gov.au/def/gsq-geochem/uom/")

VOCAB_DIR = Path(__file__).parent / "vocabs-uom"
OUTPUT_FILE = VOCAB_DIR / "uom.ttl"

wb = _load_workbook(filename="../examples/data-submission-template-minerals-3.0-001.xlsx", data_only=True)
ws = wb["UNITS_OF_MEASURE"]


def id_to_iri(id: str):
    replacements = {
        ' ': '-',
        '(': '',
        ')': '',
        '/': '-',
        '±': '',
        ',': '',
        '#': 'No',
        '"': "in",
        '.': ''
    }
    replaced_chars = [replacements.get(char, char) for char in id.lower()]
    return ''.join(replaced_chars).replace("--", "-")

concepts = set()

col = 0
while True:
    col += 1
    val = ws.cell(row=1, column=col).value

    if val is None:
        break
    else:
        print(f"Column {val}")
        collection_name = str(val).title().replace("_", " ")
        collection_iri = id_to_iri(val)

        row = 1
        while True:
            val = ws.cell(row=row, column=col).value
            if val is None:
                break
            else:
                row += 1
                val = ws.cell(row=row, column=col).value
                if val is None:
                    break
                concept_code = str(val).split("(")[1].split(")")[0]
                concepts.add((concept_code, collection_iri))

with open("vocabs-uom/uom.csv", "w") as f:
    f.write(f"ExcelCode,Collection\n")
    for concept in sorted(concepts):
        f.write(f"{concept[0]},{concept[1]}\n")
