"""This script creates a stand-alon Excel workbook called validation-dictionary.xlsx with a single whorksheet called VALIDATION_DICTIONARY that contains codlist versions of all the vocabularies in a given directory.

skos:notation values for the vocabs are used to create headers for each codelists.

skos:notation values for each Concept in the vocab are used to create the code values.

Set the directory of vocabs by modifying the VOCABS_DIR variable.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Border, PatternFill, Side
from rdflib import Graph
from rdflib.namespace import RDF, SKOS

VOCABS_DIR = Path(__file__).parent / "vocabs-codelists" / "3.0"
OUTPUT_FILE = Path(__file__).parent / "validation-dictionary.xlsx"

print(f"vocabs dir: {VOCABS_DIR}")

# styling
ft = Font(bold=True)
light_yellow = PatternFill("solid", fgColor="00FFFF99")
border_thin = Side(border_style="thin", color="000000")
border_double = Side(border_style="double", color="000000")


wb = Workbook()
ws = wb.active
ws.title = "VALIDATION_DICTIONARY"

col = 1
for f in sorted(VOCABS_DIR.glob("*.ttl")):
    print(f)
    g = Graph().parse(f)
    cs = g.value(predicate=RDF.type, object=SKOS.ConceptScheme)
    list_id = g.value(subject=cs, predicate=SKOS.notation)
    print(list_id)

    ws.cell(row=1, column=col).value = list_id

    notations = []
    for c in g.subjects(RDF.type, SKOS.Concept):
        notations.append(g.value(subject=c, predicate=SKOS.notation))

    for i, notation in enumerate(sorted(notations)):
        ws.cell(row=i+2, column=col).value = notation
        ws.cell(row=i + 2, column=col).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

    col += 1

# apply header style
for col in ws.columns:
    for cell in col:
        if cell.value is not None:
            cell.font = ft
            cell.fill = light_yellow
            cell.border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_double)
        break

wb.save(OUTPUT_FILE)
