import logging
import re
from pathlib import Path
from tempfile import SpooledTemporaryFile
from typing import Dict, Tuple, Union, Optional, List
from typing import Literal as TypeLiteral
import dateparser
import datetime

import openpyxl
import pyshacl
from colorama import Fore, Style
from openpyxl import load_workbook as _load_workbook
from openpyxl.workbook.workbook import Workbook
from pyshacl.pytypes import GraphLike
from rdflib import BNode, Graph, Literal, Namespace, URIRef
from rdflib.namespace import DCAT, DCTERMS, PROV, RDF, RDFS, SKOS, XSD, TIME
from .defined_namespaces import BORE, FOIS, SAMPLES
import utm

EXCEL_FILE_ENDINGS = ["xlsx"]
RDF_FILE_ENDINGS = {
    ".ttl": "ttl",
    ".rdf": "xml",
    ".xml": "xml",
    ".json-ld": "json-ld",
    ".json": "json-ld",
    ".nt": "nt",
    ".n3": "n3",
}
KNOWN_FILE_ENDINGS = [str(x) for x in RDF_FILE_ENDINGS.keys()] + EXCEL_FILE_ENDINGS
KNOWN_TEMPLATE_VERSIONS = [
    "3.0",
]
LATEST_TEMPLATE = KNOWN_TEMPLATE_VERSIONS[-1]
VOCAB_COLUMNS = {
    "LEASE_NAME": "A",
    "COORD_SYS_ID": "B",
    "DEPTH_DATUM": "C",
    "LOC_SURVEY_TYPE": "D",
    "DRILL_TYPE": "E",
    "DRILL_DIAMETER": "F",
    "CURRENT_CLASS": "G",
    "RPT_SURVEY_TYPE": "H",
    "SAMPLE_MATERIAL": "I",
    "SAMPLE_TYPE_DRILLING": "J",
    "SAMPLE_TYPE_SURFACE": "K",
    "MESH_SIZE": "L",
    "QAQC": "M",
    "SOIL_COLOUR": "N",
    "WEATHERING": "O",
    "ALTERATION": "P",
    "STRUCTURE": "Q",
    "TEXTURE": "R",
    "GRAIN_SIZE": "S",
    "STRUCTURAL_FEATURE": "T",
    "COMMODITY": "U",
    "RESERVE_CLASS_ID": "V",
    "RESOURCE_STATUS": "W",
    "AGENT": "X",
}
VOCABS_DIR_30 = Path(__file__).parent.parent.resolve().parent / "vocabs" / "vocabs-codelists" / "3.0"
FIELD_VOCABS = {
    "SAMPLE_MATERIAL": VOCABS_DIR_30 / "gsq-sample-materials.ttl",
    #"SAMPLE_TYPE_SURFACE": VOCABS_DIR / "sample-type-surfaces.ttl",
    "MESH_SIZE": VOCABS_DIR_30 / "sample-mesh-sizes.ttl",
    "SOIL_COLOUR": VOCABS_DIR_30 / "soil-colour.ttl",
    "AGENT": VOCABS_DIR_30 / "agents.ttl",
}

SHEETS_VOCABS = {
    "DICTIONARY": {
        "B": {
            "tab": "VALIDATION_DICTIONARY",
            "column_name": "METHOD_TYPE"
        }
    },
    "DRILLHOLE_SAMPLE": {
        "D": {
            "tab": "VALIDATION_DICTIONARY",
            "column_name": "SAMPLE_TYPE_DRILLING",
        }
    },
    "SURFACE_SAMPLE": {
        "C": {
            "tab": "VALIDATION_DICTIONARY",
            "column_name": "SAMPLE_MATERIAL",
        },
        "D": {
            "tab": "VALIDATION_DICTIONARY",
            "column_name": "SAMPLE_TYPE_SURFACE",
        },
        "E": {
            "tab": "VALIDATION_DICTIONARY",
            "column_name": "MESH_SIZE",
        },
        "G": {
            "tab": "VALIDATION_DICTIONARY",
            "column_name": "SOIL_COLOUR",
        },
        "L": {
            "tab": "VALIDATION_DICTIONARY",
            "column_name": "LOC_SURVEY_TYPE"
        },
    },
    "SAMPLE_PREPARATION": {
        # "E":
        "F": {
            "tab": "DICTIONARY",
            "column_name": "CODE"
        }
    }
}


class ConversionError(Exception):
    pass


def load_workbook(file_path: Path) -> Workbook:
    if not isinstance(
        file_path, SpooledTemporaryFile
    ) and not file_path.name.lower().endswith(tuple(EXCEL_FILE_ENDINGS)):
        raise ValueError("Files for conversion to RDF must be Excel files ending .xlsx")
    return _load_workbook(filename=file_path, data_only=False)


def load_template(file_path: Path) -> Workbook:
    if not file_path.name.lower().endswith(tuple(EXCEL_FILE_ENDINGS)):
        raise ValueError(
            "Template files for RDF-to-Excel conversion must be Excel files ending .xlsx"
        )

    return _load_workbook(filename=str(file_path), data_only=True)


def get_template_version(wb: Workbook) -> str:
    notes_sheet = wb["TEMPLATE_NOTES"]
    version_number = notes_sheet["C7"].value.replace("VERSION ", "")
    if version_number in KNOWN_TEMPLATE_VERSIONS:
        return version_number
    else:
        if float(version_number) < 2.0:
            raise Exception(
                f"You are using tempalte version {version_number} however only versions 2.0+ are supported"
            )

    raise Exception(
        "The version of the Excel template you are using cannot be determined"
    )


def split_and_tidy_to_strings(s: str):
    # note this may not work in list of things that contain commas. Need to consider revising
    # to allow comma-seperated values where it'll split in commas but not in things enclosed in quotes.
    if s == "" or s is None:
        return []
    else:
        return [x.strip() for x in re.split("[,\n]\s?", s.strip()) if x != ""]


def split_and_tidy_to_iris(s: str, prefixes):
    return [
        expand_namespaces(ss.strip(), prefixes) for ss in split_and_tidy_to_strings(s)
    ]


def string_is_http_iri(s: str) -> Tuple[bool, str]:
    # returns (True, None) if the string (sort of) is an IRI
    # returns (False, message) otherwise
    if s is None:
        return None

    messages = []
    if not s.startswith("http"):
        messages.append(
            f"HTTP IRIs must start with 'http' or 'https'. Your value was '{s}'"
        )
        if ":" in s:
            messages.append(
                f"It looks like your IRI might contain a prefix, {s.split(':')[0]+':'}, that could not be expanded. "
                "Check it's present in the Prefixes sheet of your workbook"
            )

    if " " in s:
        messages.append("IRIs cannot contain spaces")

    if len(messages) > 0:
        return False, " and ".join(messages)
    else:
        return True, ""


def all_strings_in_list_are_iris(l_: []) -> Tuple[bool, str]:
    messages = []
    if l_ is not None:
        for item in l_:
            r = string_is_http_iri(item)
            if not r[0]:
                messages.append(f"Item {item} failed with messages {r[1]}")

    if len(messages) > 0:
        return False, " and ".join(messages)
    else:
        return True, ""


def expand_namespaces(s: str, prefixes: dict[str, Namespace]) -> Union[URIRef, str]:
    for pre in prefixes.keys():
        if s.startswith(pre):
            return URIRef(s.replace(pre, prefixes[pre]))
    if s.startswith("http"):
        return URIRef(s)
    else:
        return s


def bind_namespaces(g: Graph, prefixes: dict[str, Namespace]):
    for pre, ns in prefixes.items():
        g.bind(pre.rstrip(":"), ns)


def string_from_iri(iri):
    s = str(iri.split("/")[-1].split("#")[-1])
    s = re.sub(r"(?<=[a-z])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])", " ", s)
    s = s.title()
    s = s.replace("-", " ")

    return s


def get_id_from_iri(iri):
    id = str(iri.split("/")[-1].split("#")[-1])
    return Literal(id, datatype=XSD.token)


def make_agent(agent_value, agent_role, prefixes, iri_of_subject) -> Graph:
    ag = Graph()
    iri = expand_namespaces(agent_value, prefixes)
    creator_iri_conv = string_is_http_iri(str(iri))
    if not creator_iri_conv[0]:
        iri = BNode()
    ag.add((iri, RDF.type, PROV.Agent))
    ag.add((iri, RDFS.label, Literal(string_from_iri(agent_value))))
    if agent_role in [DCTERMS.creator, DCTERMS.publisher, DCTERMS.rightsHolder]:
        ag.add((iri_of_subject, agent_role, iri))
    else:
        qa = BNode()
        ag.add((iri_of_subject, PROV.qualifiedAttribution, qa))
        ag.add((qa, PROV.agent, iri))
        ag.add((qa, DCAT.hadRole, agent_role))

    return ag


def make_iri(s: str, prefixes: dict[str, Namespace]):
    iri = expand_namespaces(s, prefixes)
    iri_conv = string_is_http_iri(str(iri))
    if not iri_conv[0]:
        raise ConversionError(iri_conv[1])
    return iri


def validate_with_profile(
    data_graph: Union[GraphLike, str, bytes],
    profile="vocpub",
    error_level=1,
    message_level=1,
    log_file=None,
):
    if profile not in profiles.PROFILES.keys():
        raise ValueError(
            "The profile chosen for conversion must be one of '{}'. 'vocpub' is default".format(
                "', '".join(profiles.PROFILES.keys())
            )
        )
    allow_warnings = True if error_level > 1 else False

    # validate the RDF file
    conforms, results_graph, results_text = pyshacl.validate(
        data_graph,
        shacl_graph=str(Path(__file__).parent / "validator.vocpub.ttl"),
        allow_warnings=allow_warnings,
    )

    logging_level = logging.INFO

    if message_level == 3:
        logging_level = logging.ERROR
    elif message_level == 2:
        logging_level = logging.WARNING

    if log_file:
        logging.basicConfig(
            level=logging_level, format="%(message)s", filename=log_file, force=True
        )
    else:
        logging.basicConfig(level=logging_level, format="%(message)s")

    info_list = []
    warning_list = []
    violation_list = []

    from rdflib.namespace import RDF, SH

    for report in results_graph.subjects(RDF.type, SH.ValidationReport):
        for result in results_graph.objects(report, SH.result):
            result_dict = {}
            for p, o in results_graph.predicate_objects(result):
                if p == SH.focusNode:
                    result_dict["focusNode"] = str(o)
                elif p == SH.resultMessage:
                    result_dict["resultMessage"] = str(o)
                elif p == SH.resultSeverity:
                    result_dict["resultSeverity"] = str(o)
                elif p == SH.sourceConstraintComponent:
                    result_dict["sourceConstraintComponent"] = str(o)
                elif p == SH.sourceShape:
                    result_dict["sourceShape"] = str(o)
                elif p == SH.value:
                    result_dict["value"] = str(o)
            result_message_formatted = log_msg(result_dict, log_file)
            result_message = log_msg(result_dict, "placeholder")
            if result_dict["resultSeverity"] == str(SH.Info):
                logging.info(result_message_formatted)
                info_list.append(result_message)
            elif result_dict["resultSeverity"] == str(SH.Warning):
                logging.warning(result_message_formatted)
                warning_list.append(result_message)
            elif result_dict["resultSeverity"] == str(SH.Violation):
                logging.error(result_message_formatted)
                violation_list.append(result_message)

    if error_level == 3:
        error_messages = violation_list
    elif error_level == 2:
        error_messages = warning_list + violation_list
    else:  # error_level == 1
        error_messages = info_list + warning_list + violation_list

    if len(error_messages) > 0:
        raise ConversionError(
            f"The file you supplied is not valid according to the {profile} profile."
        )


def log_msg(result: Dict, log_file: str) -> str:
    from rdflib.namespace import SH

    formatted_msg = ""
    message = f"""Validation Result in {result['sourceConstraintComponent'].split(str(SH))[1]} ({result['sourceConstraintComponent']}):
\tSeverity: sh:{result['resultSeverity'].split(str(SH))[1]}
\tSource Shape: <{result['sourceShape']}>
\tFocus Node: <{result['focusNode']}>
\tValue Node: <{result.get('value', '')}>
\tMessage: {result['resultMessage']}
"""
    if result["resultSeverity"] == str(SH.Info):
        formatted_msg = (
            f"INFO: {message}"
            if log_file
            else Fore.BLUE + "INFO: " + Style.RESET_ALL + message
        )
    elif result["resultSeverity"] == str(SH.Warning):
        formatted_msg = (
            f"WARNING: {message}"
            if log_file
            else Fore.YELLOW + "WARNING: " + Style.RESET_ALL + message
        )
    elif result["resultSeverity"] == str(SH.Violation):
        formatted_msg = (
            f"VIOLATION: {message}"
            if log_file
            else Fore.RED + "VIOLATION: " + Style.RESET_ALL + message
        )
    return formatted_msg


def is_a_concept_in(s: str, vocab_file: Path):
    g = Graph().parse(vocab_file)
    # try IRI & notation ID
    for iri in g.subjects(RDF.type, SKOS.Concept):
        if s == str(iri) or s == str(g.value(iri, SKOS.notation)):
            return [True]

    lbl = vocab_file.name
    for cs in g.subjects(RDF.type, SKOS.ConceptScheme):
        for label in g.objects(cs, SKOS.prefLabel):
            lbl = label

    return [False, f"IRI {s} is not a Concept in {lbl}"]


def convert_easting_northing_elevation_to_wkt(easting, northing, elevation) -> str:
    lat, lon = utm.to_latlon(easting, northing, 55, "E")

    return f"POINTZ({lon:.6f} {lat:.6f} {elevation})"


def create_vocab_validation_formula(wb: Workbook, tab, column_name: str):
    """returns something like =VALIDATION_DICTIONARY!$J$5:$J$13 """
    ws = wb[tab]
    smallest_row = 2 if tab == "VALIDATION_DICTIONARY" else 9
    greatest_row = 5
    for col in range(1, 100):
        header_cell = ws.cell(column=col, row=1)
        if header_cell.value == column_name:
            col_letter = header_cell.column_letter
            for row in ws.iter_rows(min_col=col, max_col=col, min_row=smallest_row):
                for cell in row:
                    if ws.cell(column=col, row=cell.row).value is None:
                        return f"={tab}!${col_letter}${smallest_row}:${col_letter}${greatest_row}"
                    else:
                        greatest_row = cell.row
    raise ValueError(f"A column header in sheet {tab} with value {column_name} could not be found")


def get_codelist_id_for_code(code: str, combined_concepts: Graph):
    # check the code is in the combined_concepts
    c = combined_concepts.value(predicate=SKOS.notation, object=Literal(code))
    if c is None:
        raise ConversionError(f"The code {code} is not in any codelist")

    # return the code of the ConceptScheme or Collection it is in
    cs = combined_concepts.value(subject=c, predicate=SKOS.inScheme)
    # if it's in the UoM vocab. we need he Collection ID, not the CS ID
    if cs != URIRef("https://linked.data.gov.au/def/gsq-geochem/uom"):
        return str(combined_concepts.value(subject=cs, predicate=SKOS.notation))
    else:
        col = combined_concepts.value(predicate=SKOS.member, object=c)
        return str(combined_concepts.value(subject=col, predicate=SKOS.notation))


def get_iri_from_code(code: str, combined_concepts: Graph) -> URIRef:
    """Returns an IRI for any valid value in the VALIATION_DICTIONARY, the USER_DICTIONARY, the UNITS_OF_MEASURE or the USER_UNITS_OF_MEASURE whorksheets"""

    return combined_concepts.value(predicate=SKOS.notation, object=Literal(code))


def validate_code(code: str, codelist_id: str, column_name: str, row_num: int, sheet_name: str, combined_concepts: Graph) -> None:
    try:  # so we get better error reporting further down
        codelist_id_actual = get_codelist_id_for_code(code, combined_concepts)
    except:
        codelist_id_actual = None
    if codelist_id_actual != codelist_id:
        raise ConversionError(
            f"The value {code} for {column_name} in row {row_num} on the {sheet_name} worksheet is not within "
            f"the {codelist_id} lookup list")


def check_template_version_supported(wb: openpyxl.Workbook):
    notes_sheet = wb["TEMPLATE_NOTES"]
    version_number = notes_sheet["C7"].value.replace("VERSION ", "")
    if version_number in KNOWN_TEMPLATE_VERSIONS:
        return version_number
    else:
        raise ConversionError(
            f"You are using template version {version_number} however only the following versions are supported: "
            f"{', '.join(KNOWN_TEMPLATE_VERSIONS)}"
        )


def make_rdflib_type(
    value,
    rdflib_type: TypeLiteral["URIRef", "Concept", "String", "Number", "Date", "Boolean"],
    combined_concepts: Optional[Graph] = None,
    uri_namespace: Optional[Namespace] = None,
):
    if value is None:
        return None
    elif rdflib_type == "URIRef":
        return uri_namespace[str(value)]
    elif rdflib_type == "Concept":
        return get_iri_from_code(value, combined_concepts)
    elif rdflib_type in ["String", "Number"]:
        return Literal(value)
    elif rdflib_type == "Date":
        return Literal(datetime.datetime.strftime(value, "%Y-%m-%d"), datatype=XSD.date)
    elif rdflib_type == "Boolean":
        return Literal(value, datatype=XSD.boolean)
