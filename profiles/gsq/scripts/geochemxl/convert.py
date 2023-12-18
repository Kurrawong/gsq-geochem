import argparse
import json
import sys
from typing import BinaryIO
from uuid import uuid4

from pyproj import Transformer
from rdflib import Namespace, Seq
from rdflib.namespace import GEO, SDO, SOSA

from .defined_namespaces import MININGROLES, TENEMENT, TENEMENTS, QLDBORES, QKINDS, GEOSAMPLE

EX = Namespace("http://example.com/")

from .utils import *
from .utils import check_template_version_supported

GSQ_PROFILE_DIR = Path(__file__).parent.parent.resolve().parent


def extract_sheet_dataset_metadata(
        wb: openpyxl.Workbook,
        combined_concepts: Graph,
        template_version: Optional[str] = None
) -> Tuple[Graph, URIRef]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = wb["DATASET_METADATA"]

    iri = sheet_name["B5"].value if string_is_http_iri(sheet_name["B5"].value) else "http://example.com/dataset/" + str(uuid4())
    name = sheet_name["B6"].value
    description = sheet_name["B7"].value
    date_created = sheet_name["B8"].value
    date_modified = sheet_name["B9"].value
    author = get_iri_from_code(sheet_name["B10"].value, combined_concepts)

    g = Graph(bind_namespaces="rdflib")
    v = URIRef(iri)
    g.add((v, RDF.type, SDO.Dataset))
    g.add((v, SDO.name, Literal(name)))
    g.add((v, SDO.description, Literal(description)))
    g.add((v, SDO.dateCreated, Literal(date_created, datatype=XSD.date)))
    g.add((v, SDO.dateModified, Literal(date_modified, datatype=XSD.date)))
    qa = BNode()
    g.add((v, PROV.qualifiedAttribution, qa))
    g.add((qa, PROV.agent, URIRef(author)))
    g.add((qa, PROV.hadRole, URIRef(
        "http://def.isotc211.org/iso19115/-1/2018/CitationAndResponsiblePartyInformation/code/CI_RoleCode/author")))

    return g, v


def validate_sheet_validation_dictionary(
        wb: openpyxl.Workbook,
        combined_concepts: Graph,
        template_version: Optional[str] = None
):
    if template_version is None:
        template_version = check_template_version_supported(wb)

    # load user dict if not present
    if not combined_concepts.value(subject=URIRef("http://example.com/user-defined-vocab"), predicate=RDF.type):
        combined_concepts += extract_sheet_user_dictionary(wb, combined_concepts)

    sheet = wb["VALIDATION_DICTIONARY"]

    # for every code in the VALIDATION_DICTIONARY sheet, check that either it's the notation of a Concept in the
    # combined_concepts or it's defined in the USER_DICTIONARY
    col = 1
    while True:
        codelist = sheet.cell(row=4, column=col).value

        if codelist is None:
            break
        else:
            if not combined_concepts.value(predicate=SKOS.notation, object=Literal(codelist)):
                raise ConversionError(f"Codelist {codelist} on worksheet VALIDATION_DICTIONARY is not known")

            row = 5
            while True:
                code = sheet.cell(row=row, column=col).value
                if code is None:
                    break
                else:
                    if not combined_concepts.value(predicate=SKOS.notation, object=Literal(code)):
                        raise ConversionError(f"Code {code} in codelist {codelist} on worksheet VALIDATION_DICTIONARY "
                                              f"is not known")
                row += 1

        col += 1

    allowed_codes = []
    for o in combined_concepts.objects(None, SKOS.notation):
        allowed_codes.append(str(o))


def extract_sheet_user_dictionary(
        wb: openpyxl.Workbook,
        template_version: Optional[str] = None,
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet = wb["USER_DICTIONARY"]

    row = 9
    if sheet["C9"].value == "MEGA":
        row = 10

    g = Graph()

    cs = URIRef("http://example.com/user-defined-vocab")
    g.add((cs, RDF.type, SKOS.ConceptScheme))
    g.add((cs, SKOS.prefLabel, Literal("User-defined Vocabulary")))
    g.add((cs, SKOS.notation, Literal("USER-VOC")))

    while True:
        if sheet[f"B{row}"].value is not None:
            bn = BNode()
            g.add((bn, RDF.type, SKOS.Concept))
            if sheet[f"C{row}"].value is None:
                raise ConversionError(
                    "You must supply a CODE value for each code you define in the USER_DICTIONARY sheet")
            g.add((bn, SKOS.notation, Literal(sheet[f"C{row}"].value)))
            g.add((bn, SKOS.inScheme, cs))
            if sheet[f"D{row}"].value is None:
                raise ConversionError(
                    "You must supply a DESCRIPTION value for each code you define in the USER_DICTIONARY sheet")
            g.add((bn, SKOS.definition, Literal(sheet[f"D{row}"].value)))

            row += 1
        else:
            break

    return g


def validate_sheet_uom(
        wb: openpyxl.Workbook,
        combined_concepts: Graph,
        template_version: Optional[str] = None
):
    if template_version is None:
        template_version = check_template_version_supported(wb)
    
    # load user UoM if not present
    if not combined_concepts.value(subject=URIRef("http://example.com/user-uom"), predicate=RDF.type):
        user_uom_g, user_uom_notations = extract_sheet_user_uom(wb, combined_concepts)
        combined_concepts += user_uom_g

    sheet = wb["UNITS_OF_MEASURE"]

    col = 1
    while True:
        codelist = sheet.cell(row=1, column=col).value

        if codelist is None:
            break
        else:
            if not combined_concepts.value(predicate=SKOS.notation, object=Literal(codelist)):
                raise ConversionError(f"Codelist {codelist} on worksheet UNITS_OF_MEASURE is not known")

            row = 2
            while True:
                code = sheet.cell(row=row, column=col).value
                if code is None:
                    break
                else:
                    code = code.split("(")[1].split(")")[0]
                    if not combined_concepts.value(predicate=SKOS.notation, object=Literal(code)):
                        raise ConversionError(f"Code {code} in codelist {codelist} on worksheet UNITS_OF_MEASURE "
                                              f"is not known")
                row += 1

        col += 1

    allowed_codes = []
    for o in combined_concepts.objects(None, SKOS.notation):
        allowed_codes.append(str(o))


def extract_sheet_user_uom(
        wb: openpyxl.Workbook,
        combined_concepts: Graph,
        template_version: Optional[str] = None
) -> Tuple[Graph, List]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet = wb["USER_UNITS_OF_MEASURE"]
    
    row = 9
    if sheet["C9"].value == "kg/L":
        row = 10

    g = Graph()
    notations = []

    cs = URIRef("http://example.com/user-defined-uom")
    g.add((cs, RDF.type, SKOS.ConceptScheme))
    g.add((cs, SKOS.prefLabel, Literal("User-defined Units of Measure")))
    g.add((cs, SKOS.notation, Literal("USER-UOM")))

    while True:
        if sheet[f"B{row}"].value is not None:
            bn = BNode()
            g.add((bn, RDF.type, SKOS.Concept))
            g.add((bn, SKOS.inScheme, cs))
            if sheet[f"B{row}"].value is None:
                raise ConversionError(
                    "You must select a COLLECTION value for each code you define in the USER_UNITS_OF_MEASURE sheet")
            col = combined_concepts.value(predicate=SKOS.notation, object=Literal(sheet[f"B{row}"].value))
            g.add((col, SKOS.member, bn))
            g.add((col, SKOS.inScheme, cs))
            if sheet[f"C{row}"].value is None:
                raise ConversionError(
                    "You must supply a UNIT_CODE value for each unit you define in the USER_UNITS_OF_MEASURE sheet")
            g.add((bn, SKOS.notation, Literal(sheet[f"C{row}"].value)))
            if sheet[f"C{row}"].value is None:
                raise ConversionError(
                    "You must supply a LABEL value for each code you define in the USER_UNITS_OF_MEASURE sheet")
            g.add((bn, SKOS.prefLabel, Literal(sheet[f"D{row}"].value, lang="en")))
            if sheet[f"C{row}"].value is None:
                raise ConversionError(
                    "You must supply a DEFINITION value for each code you define in the USER_UNITS_OF_MEASURE sheet")
            g.add((bn, SKOS.definition, Literal(sheet[f"E{row}"].value, lang="en")))

            notations.append(sheet[f"C{row}"].value)

            row += 1
        else:
            break

    return g, notations


def extract_sheet_user_sample_prep_codes(
        wb: openpyxl.Workbook,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Tuple[Graph, List]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "USER_SAMPLE_PREP_CODES"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()
    code_ids = []

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv in ["WEI-21x", "CRU-21x", "SPL-01x", "CRU-36fx"]:
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "code": bv,
                        "description": sheet[f"C{row}"].value,
                    },
                    "optional": {
                        "citation": sheet[f"D{row}"].value,
                    }
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                # None

                # make RDFLib objects of the values
                code_id = make_id_from_name(data["required"]["code"])
                code_lit = make_rdflib_type(data["required"]["code"], "String")
                code_iri = make_rdflib_type(code_id, "URIRef", uri_namespace=Namespace(str(dataset_iri) + "/code/"))
                description_lit = make_rdflib_type(data["required"]["description"], "String")
                if data["optional"].get("citation") is not None:
                    citation_lit = make_rdflib_type(data["optional"]["citation"], "String")
                cs_iri = URIRef(str(dataset_iri) + "/user-ConceptScheme-sample-preparations")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, cs_iri))

                g.add((cs_iri, RDF.type, SKOS.ConceptScheme))
                g.add((cs_iri, SKOS.prefLabel, Literal("User-defined Preparations", lang="en")))

                g.add((code_iri, RDF.type, SKOS.Concept))
                g.add((code_iri, SKOS.prefLabel, code_lit))
                g.add((code_iri, SKOS.definition, description_lit))
                if data["optional"].get("citation") is not None:
                    g.add((code_iri, SDO.citation, citation_lit))

                g.add((code_iri, SKOS.inScheme, cs_iri))
                g.add((code_iri, SKOS.topConceptOf, cs_iri))
                g.add((cs_iri, SKOS.hasTopConcept, code_iri))

                code_ids.append(code_id)

                row += 1
        else:
            break

    return g, code_ids


def extract_sheet_user_assay_codes(
        wb: openpyxl.Workbook,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Tuple[Graph, List]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "USER_ASSAY_CODES"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()
    code_ids = []

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv in ["IC587x", "FA50x", "BLEGx"]:
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "code": bv,
                        "description": sheet[f"C{row}"].value,
                    },
                    "optional": {
                        "citation": sheet[f"D{row}"].value,
                    }
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                # None

                # make RDFLib objects of the values
                code_id = make_id_from_name(data["required"]["code"])
                code_lit = make_rdflib_type(data["required"]["code"], "String")
                code_iri = make_rdflib_type(code_id, "URIRef", uri_namespace=Namespace(str(dataset_iri) + "/code/"))
                description_lit = make_rdflib_type(data["required"]["description"], "String")
                if data["optional"].get("citation") is not None:
                    citation_lit = make_rdflib_type(data["optional"]["citation"], "String")
                cs_iri = URIRef(str(dataset_iri) + "/user-ConceptScheme-assays")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, cs_iri))

                g.add((cs_iri, RDF.type, SKOS.ConceptScheme))
                g.add((cs_iri, SKOS.prefLabel,  Literal("User-defined Assays", lang="en")))

                g.add((code_iri, RDF.type, SKOS.Concept))
                g.add((code_iri, SKOS.prefLabel, code_lit))
                g.add((code_iri, SKOS.definition, description_lit))
                if data["optional"].get("citation") is not None:
                    g.add((code_iri, SDO.citation, citation_lit))

                g.add((code_iri, SKOS.inScheme, cs_iri))
                g.add((code_iri, SKOS.topConceptOf, cs_iri))
                g.add((cs_iri, SKOS.hasTopConcept, code_iri))

                code_ids.append(code_id)

                row += 1
        else:
            break

    return g, code_ids


def extract_sheet_user_laboratories(
        wb: openpyxl.Workbook,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Tuple[Graph, Dict]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "USER_LABORATORIES"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()
    labs_dict = {}

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv in ["GeoChem Labs Pty Ltd", "XYZ Corp (TSV)", "XYZ Corp North"]:
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "laboratory_name": bv,
                        "laboratory_location": sheet[f"C{row}"].value,
                    },
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                # None

                # make RDFLib objects of the values
                laboratory_id = make_id_from_name(data["required"]["laboratory_name"])
                laboratory_iri = make_rdflib_type(laboratory_id, "URIRef", uri_namespace=Namespace(str(dataset_iri) + "/lab/"))
                laboratory_name_lit = make_rdflib_type(data["required"]["laboratory_name"], "String")
                laboratory_location_lit = make_rdflib_type(data["required"]["laboratory_location"], "String")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, laboratory_iri))

                g.add((laboratory_iri, RDF.type, SDO.Organization))
                g.add((laboratory_iri, SDO.name, laboratory_name_lit))
                g.add((laboratory_iri, SDO.location, laboratory_location_lit))

                labs_dict[data["required"]["laboratory_name"]] = laboratory_iri

                row += 1
        else:
            break

    g.bind("ex", EX)

    return g, labs_dict


def extract_sheet_user_analytes(
        wb: openpyxl.Workbook,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Tuple[Graph, List]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "USER_ANALYTES"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()
    code_ids = []

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv in ["Au_example", "Al_example", "Bauxite_example"]:
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "code": bv,
                        "description": sheet[f"C{row}"].value,
                    },
                    "optional": {
                        "citation": sheet[f"D{row}"].value,
                    }
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                # None

                # make RDFLib objects of the values
                code_id = make_id_from_name(data["required"]["code"])
                code_lit = make_rdflib_type(data["required"]["code"], "String")
                code_iri = make_rdflib_type(code_id, "URIRef", uri_namespace=Namespace(str(dataset_iri) + "/code/"))
                description_lit = make_rdflib_type(data["required"]["description"], "String")
                if data["optional"].get("citation") is not None:
                    citation_lit = make_rdflib_type(data["optional"]["citation"], "String")
                cs_iri = URIRef(str(dataset_iri) + "/user-ConceptScheme-analytes")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, cs_iri))

                g.add((cs_iri, RDF.type, SKOS.ConceptScheme))
                g.add((cs_iri, SKOS.prefLabel,  Literal("User-defined Analytes", lang="en")))

                g.add((code_iri, RDF.type, SKOS.Concept))
                g.add((code_iri, SKOS.prefLabel, code_lit))
                g.add((code_iri, SKOS.definition, description_lit))
                if data["optional"].get("citation") is not None:
                    g.add((code_iri, SDO.citation, citation_lit))

                g.add((code_iri, SKOS.inScheme, cs_iri))
                g.add((code_iri, SKOS.topConceptOf, cs_iri))
                g.add((cs_iri, SKOS.hasTopConcept, code_iri))

                code_ids.append(code_id)

                row += 1
        else:
            break

    return g, code_ids


def extract_sheet_tenement(
        wb: openpyxl.Workbook,
        combined_concepts: Graph,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "TENEMENT"
    sheet = wb[sheet_name]

    row = 9
    if sheet["C9"].value == 12345:
        row = 10

    g = Graph()

    while True:
        if sheet[f"B{row}"].value is not None:
            # make vars of all the sheet values
            data = {
                "required": {
                    "tenement_type": sheet[f"B{row}"].value,
                    "tenement_no": sheet[f"C{row}"].value,
                    "tenement_holder": sheet[f"D{row}"].value,
                    "project_name": sheet[f"E{row}"].value,
                    "tenement_operator": sheet[f"F{row}"].value,
                    "geodetic_datum": sheet[f"G{row}"].value,
                    "map_sheet_no": sheet[f"H{row}"].value,
                },
                "optional": {
                    "remark": sheet[f"I{row}"].value,
                }
            }

            # check required sheet values are present
            for k, v in data["required"].items():
                if v is None:
                    raise ConversionError(
                        f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

            # check lookup values are valid
            validate_code(
                data["required"]["tenement_type"], "LEASE_NAME", "TENEMENT_TYPE", row, sheet_name, combined_concepts
            )

            validate_code(
                data["required"]["geodetic_datum"], "COORD_SYS_ID", "GEODETIC_DATUM", row, sheet_name, combined_concepts
            )

            # make RDFLib objects of the values
            tenement_iri = URIRef(TENEMENTS + str(data["required"]["tenement_no"]))
            tenement_type_iri = get_iri_from_code(data["required"]["tenement_type"], combined_concepts)
            tenement_holder_lit = Literal(data["required"]["tenement_holder"])
            project_name_lit = Literal(data["required"]["project_name"])
            tenement_operator_lit = Literal(data["required"]["tenement_operator"])
            geodetic_datum_iri = get_iri_from_code(data["required"]["geodetic_datum"], combined_concepts)
            map_sheet_no_lit = [
                Literal(x.strip(), datatype=TENEMENT.MapSheet)
                for x in str(data["required"]["map_sheet_no"]).split(",")
            ]

            remark_lit = Literal(data["optional"]["remark"])

            # make the graph
            g.add((dataset_iri, SDO.hasPart, tenement_iri))

            g.add((tenement_iri, RDF.type, TENEMENT.Tenement))

            g.add((tenement_iri, SDO.additionalType, tenement_type_iri))

            qa = BNode()
            g.add((tenement_iri, PROV.qualifiedAttribution, qa))
            g.add((qa, PROV.agent, tenement_holder_lit))
            g.add((qa, PROV.hadRole, MININGROLES.TenementHolder))

            g.add((tenement_iri, TENEMENT.hasProject, project_name_lit))

            qa2 = BNode()
            g.add((tenement_iri, PROV.qualifiedAttribution, qa2))
            g.add((qa2, PROV.agent, tenement_operator_lit))
            g.add((qa2, PROV.hadRole, MININGROLES.TenementOperator))

            ta = BNode()
            g.add((ta, RDF.type, GEO.Feature))
            g.add((ta, RDF.type, TENEMENT.TenementArea))

            g.add((tenement_iri, SDO.location, ta))

            tg = BNode()
            g.add((tg, RDF.type, GEO.Geometry))
            g.add((tg, RDFS.comment, Literal(f"CRS is {geodetic_datum_iri}")))
            for map_sheet in map_sheet_no_lit:
                g.add((tg, SDO.identifier, map_sheet))

            g.add((ta, GEO.hasGeometry, tg))

            if data["optional"]["remark"] is not None:
                g.add((tenement_iri, RDFS.comment, remark_lit))

            row += 1
        else:
            break

    g.bind(TENEMENT.prefix, TENEMENT)

    return g


def extract_sheet_drillhole_location(
        wb: openpyxl.Workbook,
        combined_concepts: Graph,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Tuple[Graph, List[str]]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "DRILLHOLE_LOCATION"
    sheet = wb[sheet_name]

    row = 9
    if sheet["B9"].value == "DD12345":
        row = 10

    g = Graph()

    drillhole_ids = []

    while True:
        if sheet[f"B{row}"].value is not None:
            # make vars of all the sheet values
            data = {
                "required": {
                    "drillhole_id": sheet[f"B{row}"].value,
                    "easting": int(sheet[f"C{row}"].value),
                    "northing": int(sheet[f"D{row}"].value),
                    "elevation": float(sheet[f"E{row}"].value),
                    "total_depth": float(sheet[f"F{row}"].value),
                    "drill_type": sheet[f"H{row}"].value,
                    "drill_diameter": sheet[f"I{row}"].value,
                    "dip": int(sheet[f"J{row}"].value),
                    "azimuth": int(sheet[f"K{row}"].value),
                    "drill_start_date": sheet[f"M{row}"].value,
                    "drill_end_date": sheet[f"N{row}"].value,
                    "location_survey_type": sheet[f"O{row}"].value,
                    "pre_collar_method": sheet[f"Q{row}"].value,
                    "pre_collar_depth": sheet[f"R{row}"].value,
                    "drill_contractor": sheet[f"S{row}"].value,
                },
                "optional": {
                    "total_depth_logger": sheet[f"G{row}"].value,
                    "current_class": sheet[f"L{row}"].value,
                    "survey_company": sheet[f"P{row}"].value,
                    "remark": sheet[f"T{row}"].value,
                }
            }

            # check required sheet values are present
            for k, v in data["required"].items():
                if v is None:
                    raise ConversionError(
                        f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

            # check lookup values are valid
            drillhole_id = data["required"]["drillhole_id"]
            drillhole_ids.append(drillhole_id)

            validate_code(
                data["required"]["drill_type"], "DRILL_TYPE", "DRILL_TYPE", row, sheet_name,
                combined_concepts
            )

            validate_code(
                data["required"]["drill_diameter"], "DRILL_DIAMETER", "DRILL_DIAMETER", row, sheet_name,
                combined_concepts
            )

            if data["optional"]["current_class"] is not None:
                validate_code(
                    data["optional"]["current_class"], "CURRENT_CLASS", "CURRENT_CLASS", row, sheet_name,
                    combined_concepts
                )

            validate_code(
                data["required"]["location_survey_type"], "LOC_SURVEY_TYPE", "LOCATION_SURVEY_TYPE", row, sheet_name,
                combined_concepts
            )

            validate_code(
                data["required"]["pre_collar_method"], "DRILL_TYPE", "PRE_COLLAR_METHOD", row, sheet_name,
                combined_concepts
            )

            # value validation
            easting = data["required"]["easting"]
            if type(easting) != int or easting < 0:
                raise ConversionError(
                    f"The value {easting} for EASTING in row {row} of sheet {sheet_name} is not an integer greater than 0"
                    f" as required")

            northing = data["required"]["northing"]
            if type(easting) != int or easting < 0:
                raise ConversionError(
                    f"The value {northing} for NORTHING in row {row} of sheet {sheet_name} is not an integer "
                    f"greater than 0 as required")

            elevation = data["required"]["elevation"]
            if type(elevation) not in [float, int]:
                raise ConversionError(
                    f"The value {elevation} for ELEVATION in row {row} of sheet {sheet_name} is not an number"
                    f" as required")

            total_depth = data["required"]["total_depth"]
            if type(total_depth) not in [float, int] and total_depth < 0:
                raise ConversionError(
                    f"The value {total_depth} for TOTAL_DEPTH in row {row} of sheet {sheet_name} is not an number"
                    f" as required")

            total_depth_logger = data["optional"].get("total_depth_logger")

            dip = data["required"]["dip"]
            if not 0 >= dip >= -90:
                raise ConversionError(
                    f"The value {dip} for DIP in row {row} of sheet {sheet_name} is not between 0 and -90 as required")

            azimuth = data["required"]["azimuth"]
            if not 0 <= azimuth <= 360:
                raise ConversionError(
                    f"The value {azimuth} for AZIMUTH in row {row} of sheet {sheet_name} is not between "
                    f"0 and 360 as required")

            drill_type = data["required"]["drill_type"]
            drill_diameter = data["required"]["drill_diameter"]

            dip = data["required"]["dip"]
            azimuth = data["required"]["azimuth"]

            current_class = data["optional"]["current_class"]

            drill_start_date = data["required"]["drill_start_date"]
            if type(drill_start_date) != datetime.datetime:
                raise ConversionError(
                    f"The value {drill_start_date} for DRILL_START_DATE in row {row} of sheet {sheet_name} "
                    f"is not a date as required")

            drill_end_date = data["required"]["drill_end_date"]
            if type(drill_end_date) != datetime.datetime:
                raise ConversionError(
                    f"The value {drill_end_date} for DRILL_END_DATE in row {row} of sheet {sheet_name} "
                    f"is not a date as required")

            location_survey_type = data["required"]["location_survey_type"]
            survey_company = data["optional"].get("survey_company")
            pre_collar_method = data["required"]["pre_collar_method"]
            pre_collar_depth = data["required"]["pre_collar_depth"]
            if type(pre_collar_depth) not in [float, int] and total_depth < 0:
                raise ConversionError(
                    f"The value {pre_collar_depth} for PRE_COLLAR_DEPTH in row {row} of sheet {sheet_name} "
                    f"is not an number as required")
            drill_contractor = data["required"]["drill_contractor"]
            remark = data["optional"].get("remark")

            # make RDFLib objects of the values
            drillhole_iri = URIRef(QLDBORES + drillhole_id)

            transformer = Transformer.from_crs("EPSG:32755", "EPSG:4326")
            lon, lat = transformer.transform(easting, northing)
            wkt = Literal(f"POINTZ({lon} {lat} {elevation})", datatype=GEO.wktLiteral)

            total_depth_lit = make_rdflib_type(total_depth, "Number")
            if total_depth_logger is not None:
                total_depth_logger_lit = make_rdflib_type(total_depth_logger, "String")

            drill_type_iri = make_rdflib_type(drill_type, "Concept", combined_concepts)
            drill_diameter_iri = make_rdflib_type(drill_diameter, "Concept", combined_concepts)

            dip_lit = make_rdflib_type(dip, "Number")
            azimuth_lit = make_rdflib_type(azimuth, "Number")

            if current_class is not None:
                current_class_iri = make_rdflib_type(current_class, "Concept", combined_concepts)

            drill_start_date_lit = make_rdflib_type(drill_start_date, "Date")
            drill_end_date_lit = make_rdflib_type(drill_end_date, "Date")
            location_survey_type_iri = make_rdflib_type(location_survey_type, "Concept", combined_concepts)
            if data["optional"]["survey_company"] is not None:
                survey_company_lit = make_rdflib_type(survey_company, "String")
            pre_collar_method_iri = make_rdflib_type(pre_collar_method, "Concept", combined_concepts)
            pre_collar_depth_lit = make_rdflib_type(pre_collar_depth, "String")
            drill_contractor_lit = make_rdflib_type(drill_contractor, "String")
            if remark is not None:
                remark_lit = make_rdflib_type(remark, "String")

            # make the graph
            g.add((dataset_iri, SDO.hasPart, drillhole_iri))

            g.add((drillhole_iri, RDF.type, BORE.Bore))

            geom = BNode()
            g.add((drillhole_iri, GEO.hasGeometry, geom))
            g.add((geom, RDF.type, GEO.Geometry))
            g.add((geom, GEO.asWKT, wkt))

            g.add((drillhole_iri, SDO.depth, total_depth_lit))

            if data["optional"]["total_depth_logger"] is not None:
                g.add((drillhole_iri, BORE.totalDepthLogger, total_depth_logger_lit))

            g.add((drillhole_iri, BORE.hadDrillingMethod, drill_type_iri))
            g.add((drillhole_iri, BORE.hasDiameter, drill_diameter_iri))
            g.add((drillhole_iri, BORE.hasCollarDip, dip_lit))
            g.add((drillhole_iri, BORE.hasCollarAzimuth, azimuth_lit))

            if data["optional"]["current_class"] is not None:
                g.add((drillhole_iri, BORE.hasPurpose, current_class_iri))

            dt = BNode()
            g.add((dt, RDF.type, BORE.DrillingTime))
            g.add((dt, PROV.startedAtTime, drill_start_date_lit))
            g.add((dt, PROV.endedAtTime, drill_end_date_lit))
            g.add((drillhole_iri, TIME.hasTime, dt))

            g.add((drillhole_iri, EX.locationSurveyType, location_survey_type_iri))

            if data["optional"]["survey_company"] is not None:
                sc = BNode()
                g.add((drillhole_iri, PROV.qualifiedAttribution, sc))
                g.add((sc, PROV.agent, survey_company_lit))
                g.add((sc, PROV.hadRole, MININGROLES.Surveyer))

            g.add((drillhole_iri, EX.preCollarMethod, pre_collar_method_iri))

            g.add((drillhole_iri, EX.preCollarDepth, pre_collar_depth_lit))

            dc = BNode()
            g.add((drillhole_iri, PROV.qualifiedAttribution, dc))
            g.add((dc, PROV.agent, drill_contractor_lit))
            g.add((dc, PROV.hadRole, MININGROLES.Driller))

            if data["optional"]["remark"] is not None:
                g.add((drillhole_iri, RDFS.comment, remark_lit))

            row += 1
        else:
            break

    g.bind("bore", BORE)
    g.bind("ex", EX)

    return g, drillhole_ids


# dependent on extract_sheet_drillhole_location
def extract_sheet_drillhole_survey(
        wb: openpyxl.Workbook,
        combined_concepts: Graph,
        drillhole_ids: List[str],
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "DRILLHOLE_SURVEY"
    sheet = wb[sheet_name]

    row = 9
    if sheet["B9"].value == "DD1234":
        row = 10

    g = Graph()

    while True:
        if sheet[f"B{row}"].value is not None:
            # make vars of all the sheet values
            data = {
                "required": {
                    "drillhole_id": sheet[f"B{row}"].value,
                    "survey_instrument": sheet[f"C{row}"].value,  # RPT_SURVEY_TYPE
                    "survey_depth": sheet[f"F{row}"].value,
                    "azimuth": sheet[f"G{row}"].value,
                    "dip": sheet[f"I{row}"].value,
                },
                "optional": {
                    "survey_company": sheet[f"D{row}"].value,
                    "survey_date": sheet[f"E{row}"].value,
                    "azimuth_accuracy": sheet[f"H{row}"].value,
                    "inclination_accuracy": sheet[f"J{row}"].value,
                    "magnetic_field": sheet[f"K{row}"].value,
                    "remark": sheet[f"L{row}"].value,
                }
            }

            # check required sheet values are present
            for k, v in data["required"].items():
                if v is None:
                    raise ConversionError(
                        f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

            # check lookup values are valid
            validate_code(
                data["required"]["survey_instrument"], "RPT_SURVEY_TYPE", "SURVEY_INSTRUMENT", row, sheet_name,
                combined_concepts
            )

            # value validation
            survey_depth = data["required"]["survey_depth"]
            if type(survey_depth) not in [float, int] and survey_depth < 0:
                raise ConversionError(
                    f"The value {survey_depth} for TOTAL_DEPTH in row {row} of sheet {sheet_name} is not an number"
                    f" as required")

            azimuth = data["required"]["azimuth"]
            if not 0 <= azimuth <= 360:
                raise ConversionError(
                    f"The value {azimuth} for AZIMUTH in row {row} of sheet {sheet_name} is not between "
                    f"0 and 360 as required")

            dip = data["required"]["dip"]
            if not 0 >= dip >= -90:
                raise ConversionError(
                    f"The value {dip} for DIP in row {row} of sheet {sheet_name} is not between 0 and -90 as required")

            survey_date = data["optional"]["survey_date"]
            if type(survey_date) != datetime.datetime:
                raise ConversionError(
                    f"The value {survey_date} for SURVEY_DATE in row {row} of sheet {sheet_name} "
                    f"is not a date as required")

            azimuth_accuracy = data["optional"]["azimuth_accuracy"]
            if azimuth_accuracy is not None:
                if not 0 < azimuth_accuracy < 100:
                    raise ConversionError(
                        f"The value {azimuth_accuracy} for DRILL_END_DATE in row {row} of sheet {sheet_name} "
                        f"is not between 0 and 100 as required")

            inclination_accuracy = data["optional"]["azimuth_accuracy"]
            if inclination_accuracy is not None:
                if not 0 < inclination_accuracy < 100:
                    raise ConversionError(
                        f"The value {inclination_accuracy} for DRILL_END_DATE in row {row} of sheet {sheet_name} "
                        f"is not between 0 and 100 as required")

            magnetic_field = data["optional"]["magnetic_field"]
            if magnetic_field is not None:
                if not 0 < magnetic_field < 10000000:
                    raise ConversionError(
                        f"The value {magnetic_field} for DRILL_END_DATE in row {row} of sheet {sheet_name} "
                        f"is not between 0 and 10000000 as required")

            remark = data["optional"].get("remark")

            # cross-sheet validation
            drillhole_id = str(data["required"]["drillhole_id"])
            if drillhole_id not in drillhole_ids:
                raise ConversionError(
                    f"The value {drillhole_id} for DRILLHOLE_ID in row {row} of sheet {sheet_name} "
                    f"is not present on sheet DRILLHOLE_LOCATION in the DRILLHOLE_ID column, as required")

            # make RDFLib objects of the values
            drillhole_iri = QLDBORES[drillhole_id]
            survey_instrument_iri = get_iri_from_code(data["required"]["survey_instrument"], combined_concepts)
            if data["optional"]["survey_company"] is not None:
                survey_company_lit = Literal(data["optional"]["survey_company"])
            if data["optional"]["survey_date"] is not None:
                survey_date_lit = Literal(datetime.datetime.strftime(data["optional"]["survey_date"], "%Y-%m-%d"), datatype=XSD.date)
            survey_depth_lit = Literal(data["required"]["survey_depth"])
            azimuth_lit = Literal(azimuth)
            if data["optional"]["azimuth_accuracy"] is not None:
                azimuth_accuracy_lit = Literal(data["optional"]["azimuth_accuracy"])
            dip_lit = Literal(dip)
            if data["optional"]["inclination_accuracy"] is not None:
                inclination_accuracy_lit = Literal(data["optional"]["inclination_accuracy"])
            if data["optional"]["magnetic_field"] is not None:
                magnetic_field_lit = Literal(data["optional"]["magnetic_field"])
            if remark is not None:
                remark_lit = Literal(data["optional"]["remark"])

            # make the graph
            g.add((dataset_iri, SDO.hasPart, drillhole_iri))

            g.add((drillhole_iri, RDF.type, BORE.Bore))
            s = BNode()
            g.add((drillhole_iri, BORE.hadSurvey, s))
            g.add((s, RDF.type, BORE.Survey))  # an ObservationCollection
            g.add((s, SOSA.madeBySensor, survey_instrument_iri))
            if data["optional"]["survey_company"] is not None:
                sc = BNode()
                g.add((s, PROV.qualifiedAttribution, sc))
                g.add((sc, PROV.agent, survey_company_lit))
                g.add((sc, PROV.hadRole, MININGROLES.Surveyer))
            if data["optional"]["survey_date"] is not None:
                t = BNode()
                g.add((t, RDF.type, TIME.Instant))
                g.add((t, TIME.inXSDDateTime, survey_date_lit))
                g.add((s, TIME.hasTime, t))

            depth_obs = BNode()
            depth_res = BNode()
            g.add((s, SOSA.hasMember, depth_obs))
            g.add((depth_obs, SOSA.observedProperty, BORE.hasTotalDepth))
            g.add((depth_obs, SOSA.hasFeatureOfInterest, drillhole_iri))
            g.add((drillhole_iri, SOSA.isFeatureOfInterestOf, depth_obs))
            g.add((depth_obs, SOSA.hasResult, depth_res))
            g.add((depth_res, SDO.value, survey_depth_lit))
            g.add((depth_res, SDO.unitCode, URIRef("http://qudt.org/vocab/unit/M")))

            az_obs = BNode()
            az_res = BNode()
            g.add((s, SOSA.hasMember, az_obs))
            g.add((az_obs, SOSA.observedProperty, BORE.hasAzimuth))
            g.add((az_obs, SOSA.hasFeatureOfInterest, drillhole_iri))
            g.add((drillhole_iri, SOSA.isFeatureOfInterestOf, az_obs))
            g.add((az_obs, SOSA.hasResult, az_res))
            g.add((az_res, SDO.value, azimuth_lit))
            g.add((az_res, SDO.unitCode, URIRef("http://qudt.org/vocab/unit/DEG")))

            if azimuth_accuracy is not None:
                g.add((az_res, SDO.marginOfError, azimuth_accuracy_lit))

            dip_obs = BNode()
            dip_res = BNode()
            g.add((s, SOSA.hasMember, dip_obs))
            g.add((dip_obs, SOSA.observedProperty, BORE.hasDip))
            g.add((dip_obs, SOSA.hasFeatureOfInterest, drillhole_iri))
            g.add((drillhole_iri, SOSA.isFeatureOfInterestOf, dip_obs))
            g.add((dip_obs, SOSA.hasResult, dip_res))
            g.add((dip_res, SDO.value, dip_lit))
            g.add((dip_res, SDO.unitCode, URIRef("http://qudt.org/vocab/unit/DEG")))

            if inclination_accuracy is not None:
                g.add((dip_res, SDO.marginOfError, inclination_accuracy_lit))

            if magnetic_field is not None:
                mag_obs = BNode()
                mag_res = BNode()
                g.add((s, SOSA.hasMember, mag_obs))
                g.add((mag_obs, SOSA.observedProperty, EX.hasMagneticFieldStrength))
                g.add((mag_obs, SOSA.hasFeatureOfInterest, drillhole_iri))
                g.add((mag_obs, SOSA.hasResult, mag_res))
                g.add((mag_res, SDO.value, magnetic_field_lit))
                g.add((mag_res, SDO.unitCode, URIRef("http://qudt.org/vocab/unit/NanoT")))

            if data["optional"]["remark"] is not None:
                g.add((s, RDFS.comment, remark_lit))

            row += 1
        else:
            break

    g.bind("bore", BORE)
    g.bind("unit", Namespace("http://qudt.org/vocab/unit/"))

    return g


# dependent on extract_sheet_drillhole_location
def extract_sheet_drillhole_sample(
        wb: openpyxl.Workbook,
        combined_concepts: Graph,
        drillhole_ids: List[str],
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "DRILLHOLE_SAMPLE"
    sheet = wb[sheet_name]

    row = 9
    if sheet["B9"].value == "DD12345":
        row = 10

    g = Graph()
    sample_ids = []

    while True:
        if sheet[f"B{row}"].value is not None:
            # make vars of all the sheet values
            data = {
                "required": {
                    "drillhole_id": sheet[f"B{row}"].value,
                    "sample_id": sheet[f"C{row}"].value,
                    "sample_type_drilling": sheet[f"D{row}"].value,  # SAMPLE_TYPE_DRILLING
                    "from": sheet[f"E{row}"].value,
                    "to": sheet[f"F{row}"].value,
                    "collection_date": sheet[f"G{row}"].value,
                    "dispatch_date": sheet[f"H{row}"].value,
                },
                "optional": {
                    "instrument_type": sheet[f"I{row}"].value,
                    "specific_gravity": sheet[f"J{row}"].value,
                    "magnetic_susceptibility": sheet[f"K{row}"].value,
                    "remark": sheet[f"L{row}"].value,
                }
            }

            # check required sheet values are present
            for k, v in data["required"].items():
                if v is None:
                    raise ConversionError(
                        f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

            # check lookup values are valid
            validate_code(
                data["required"]["sample_type_drilling"], "SAMPLE_TYPE_DRILLING", "SAMPLE_TYPE_DRILLING", row, sheet_name,
                combined_concepts
            )

            # value validation
            depth_from = data["required"]["from"]
            if depth_from < 0:
                raise ConversionError(
                    f"The value {depth_from} for FROM in row {row} of sheet {sheet_name} is not greater or equal to "
                    f"zero as required")

            depth_to = data["required"]["to"]
            if depth_to <= depth_from:
                raise ConversionError(
                    f"The value {depth_to} for TO in row {row} of sheet {sheet_name} is not greater or equal to "
                    f"the FROM value as required")

            collection_date = data["required"]["collection_date"]
            if type(collection_date) != datetime.datetime:
                collection_date = dateparser.parse(collection_date)
                if type(collection_date) != datetime.datetime:
                    raise ConversionError(
                        f'The value {data["required"]["collection_date"]} for COLLECTION_DATE in row {row} of '
                        f'sheet {sheet_name} is not a date as required')

            dispatch_date = data["required"]["dispatch_date"]
            if type(dispatch_date) != datetime.datetime:
                dispatch_date = dateparser.parse(dispatch_date)
                if type(dispatch_date) != datetime.datetime:
                    raise ConversionError(
                        f'The value {data["required"]["dispatch_date"]} for DISPATCH_DATE in row {row} of '
                        f'sheet {sheet_name} is not a date as required')

            if dispatch_date < collection_date:
                raise ConversionError(
                    f"The value {dispatch_date} for DISPATCH_DATE in row {row} of sheet {sheet_name} "
                    f"is not greater than or equal to the value {dispatch_date} in the same row, as required")

            instrument_type = data["optional"].get("instrument_type")

            specific_gravity = data["optional"].get("specific_gravity")
            if specific_gravity is not None:
                if specific_gravity < 0:
                    raise ConversionError(
                        f"The value {specific_gravity} for SPECIFIC_GRAVITY in row {row} of sheet {sheet_name} "
                        f"is not greater than 0, as required")

            magnetic_susceptibility = data["optional"].get("magnetic_susceptibility")
            if magnetic_susceptibility is not None:
                if not str(magnetic_susceptibility).startswith("-"):
                    raise ConversionError(
                        f"The value {magnetic_susceptibility} for MAGNETIC_SUSCEPTIBILITY in row {row} of sheet {sheet_name} "
                        f"is not negative, as required")

            remark = data["optional"].get("remark")

            # cross-sheet validation
            drillhole_id = str(data["required"]["drillhole_id"])
            if drillhole_id not in drillhole_ids:
                raise ConversionError(
                    f"The value {drillhole_id} for DRILLHOLE_ID in row {row} of sheet {sheet_name} "
                    f"is not present on sheet DRILLHOLE_LOCATION in the DRILLHOLE_ID column, as required")

            # make RDFLib objects of the values
            drillhole_iri = QLDBORES[drillhole_id]
            sample_iri = make_rdflib_type(data["required"]["sample_id"], "URIRef", None, Namespace(dataset_iri + "/sample/"))
            sample_type_drilling_iri = make_rdflib_type(data["required"]["sample_type_drilling"], "Concept", combined_concepts)
            depth_from_lit = make_rdflib_type(depth_from, "Number")
            depth_to_lit = make_rdflib_type(depth_to, "Number")
            collection_date_lit = make_rdflib_type(collection_date, "Date")
            dispatch_date_lit = make_rdflib_type(dispatch_date, "Date")

            if instrument_type is not None:
                instrument_type_lit = make_rdflib_type(instrument_type, "String")
            if specific_gravity is not None:
                specific_gravity_lit = make_rdflib_type(specific_gravity, "Number")
            if magnetic_susceptibility is not None:
                magnetic_susceptibility_lit = make_rdflib_type(magnetic_susceptibility, "Number")
            if remark is not None:
                remark_lit = make_rdflib_type(remark, "String")

            # make the graph
            g.add((dataset_iri, SDO.hasPart, drillhole_iri))
            g.add((dataset_iri, SDO.hasPart, sample_iri))

            g.add((drillhole_iri, RDF.type, BORE.Bore))
            g.add((sample_iri, RDF.type, SOSA.Sample))
            g.add((sample_iri, SOSA.isSampleOf, drillhole_iri))
            g.add((sample_iri, SDO.additionalType, URIRef("ttps://linked.data.gov.au/def/sample-type/rock")))
            g.add((sample_iri, SOSA.usedProcedure, sample_type_drilling_iri))
            g.add((sample_iri, SDO.depth, depth_from_lit))
            g.add((sample_iri, SDO.depth, depth_to_lit))
            g.add((sample_iri, PROV.generatedAtTime, collection_date_lit))
            g.add((sample_iri, SDO.dateIssued, dispatch_date_lit))
            if instrument_type is not None:
                g.add((sample_iri, SOSA.madeBySensor, instrument_type_lit))

            if specific_gravity is not None:
                spec_grav_obs = BNode()
                spec_grav_res = BNode()
                g.add((spec_grav_obs, RDF.type, SOSA.Observation))
                g.add((spec_grav_obs, SOSA.hasFeatureOfInterest, sample_iri))
                g.add((sample_iri, SOSA.isFeatureOfInterestOf, spec_grav_obs))
                g.add((spec_grav_obs, SOSA.observedProperty, EX.SpecificGravity))
                g.add((spec_grav_obs, SOSA.hasResult, spec_grav_res))
                g.add((spec_grav_res, RDF.type, SOSA.Result))
                g.add((spec_grav_res, SDO.value, specific_gravity_lit))
            if magnetic_susceptibility is not None:
                mag_sup_obs = BNode()
                mag_sup_res = BNode()
                g.add((mag_sup_obs, RDF.type, SOSA.Observation))
                g.add((mag_sup_obs, SOSA.hasFeatureOfInterest, sample_iri))
                g.add((sample_iri, SOSA.isFeatureOfInterestOf, mag_sup_obs))
                g.add((mag_sup_obs, SOSA.observedProperty, QKINDS.MagneticSusceptability))
                g.add((mag_sup_obs, SOSA.hasResult, mag_sup_res))
                g.add((mag_sup_res, RDF.type, SOSA.Result))
                g.add((mag_sup_res, SDO.value, magnetic_susceptibility_lit))
            if remark is not None:
                g.add((sample_iri, RDFS.comment, remark_lit))

            sample_ids.append(data["required"]["sample_id"])

            row += 1
        else:
            break

    g.bind("bore", BORE)
    g.bind("qkinds", QKINDS)

    return g, sample_ids


def extract_sheet_surface_sample(
        wb: openpyxl.Workbook,
        combined_concepts: Graph,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "SURFACE_SAMPLE"
    sheet = wb[sheet_name]

    row = 9
    if sheet["B9"].value == "SS12345":
        row = 10
    if sheet["B10"].value == "SS12346":
        row = 11
    if sheet["B11"].value == "SS12347":
        row = 12

    g = Graph()
    sample_ids = []

    while True:
        if sheet[f"B{row}"].value is not None:
            # make vars of all the sheet values
            data = {
                "required": {
                    "sample_id": sheet[f"B{row}"].value,
                    "sample_material": sheet[f"C{row}"].value,  # SAMPLE_MATERIAL
                    "sample_type_surface": sheet[f"D{row}"].value,  # SAMPLE_TYPE_SURFACE
                    "sample_mesh_size": sheet[f"E{row}"].value,  # MESH_SIZE
                    "soil_sample_depth": sheet[f"F{row}"].value,
                    "soil_colour": sheet[f"G{row}"].value,  # COLOUR
                    "soil_ph": sheet[f"H{row}"].value,
                    "easting": sheet[f"I{row}"].value,
                    "northing": sheet[f"J{row}"].value,
                    "location_survey_type": sheet[f"L{row}"].value,
                    "collection_date": sheet[f"M{row}"].value,
                    "dispatch_date": sheet[f"N{row}"].value,
                },
                "optional": {
                    "elevation": sheet[f"K{row}"].value,
                    "instrument_type": sheet[f"O{row}"].value,
                    "specific_gravity": sheet[f"P{row}"].value,
                    "magnetic_susceptibility": sheet[f"Q{row}"].value,
                    "remark": sheet[f"R{row}"].value,
                }
            }

            # check required sheet values are present
            for k, v in data["required"].items():
                if v is None:
                    raise ConversionError(
                        f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

            # check lookup values are valid
            validate_code(
                data["required"]["sample_material"], "SAMPLE_MATERIAL", "SAMPLE_MATERIAL", row,
                sheet_name,
                combined_concepts
            )

            validate_code(
                data["required"]["sample_type_surface"], "SAMPLE_TYPE_SURFACE", "SAMPLE_TYPE_SURFACE", row,
                sheet_name,
                combined_concepts
            )

            validate_code(
                data["required"]["sample_mesh_size"], "MESH_SIZE", "MESH_SIZE", row,
                sheet_name,
                combined_concepts
            )

            validate_code(
                data["required"]["soil_colour"], "COLOUR", "SOIL_COLOUR", row,
                sheet_name,
                combined_concepts
            )

            validate_code(
                data["required"]["location_survey_type"], "LOC_SURVEY_TYPE", "LOCATION_SURVEY_TYPE", row, sheet_name,
                combined_concepts
            )

            # value validation
            soil_sample_depth = data["required"]["soil_sample_depth"]
            if soil_sample_depth < 0:
                raise ConversionError(
                    f"The value {soil_sample_depth} for SOIL_SAMPLE_DEPTH in row {row} of sheet {sheet_name} is "
                    f"not greater or equal to zero as required")

            soil_ph = data["required"]["soil_ph"]
            if soil_ph < 0 or soil_ph > 14:
                raise ConversionError(
                    f"The value {soil_sample_depth} for SOIL_PH in row {row} of sheet {sheet_name} is "
                    f"is not between 0 and 14 as required")

            easting = data["required"]["easting"]
            if type(easting) != int or easting < 0:
                raise ConversionError(
                    f"The value {easting} for EASTING in row {row} of sheet {sheet_name} is not an integer greater than 0"
                    f" as required")

            northing = data["required"]["northing"]
            if type(easting) != int or easting < 0:
                raise ConversionError(
                    f"The value {northing} for NORTHING in row {row} of sheet {sheet_name} is not an integer "
                    f"greater than 0 as required")

            elevation = data["optional"]["elevation"]
            if elevation is not None:
                if type(elevation) not in [float, int]:
                    raise ConversionError(
                        f"The value {elevation} for ELEVATION in row {row} of sheet {sheet_name} is not an number"
                        f" as required")

            collection_date = data["required"]["collection_date"]
            if type(collection_date) != datetime.datetime:
                collection_date = dateparser.parse(collection_date)
                if type(collection_date) != datetime.datetime:
                    raise ConversionError(
                        f'The value {data["required"]["collection_date"]} for COLLECTION_DATE in row {row} of '
                        f'sheet {sheet_name} is not a date as required')

            dispatch_date = data["required"]["dispatch_date"]
            if type(dispatch_date) != datetime.datetime:
                dispatch_date = dateparser.parse(dispatch_date)
                if type(dispatch_date) != datetime.datetime:
                    raise ConversionError(
                        f'The value {data["required"]["dispatch_date"]} for DISPATCH_DATE in row {row} of '
                        f'sheet {sheet_name} is not a date as required')

            if dispatch_date < collection_date:
                raise ConversionError(
                    f"The value {dispatch_date} for DISPATCH_DATE in row {row} of sheet {sheet_name} "
                    f"is not greater than or equal to the value {collection_date} in the same row, as required")

            instrument_type = data["optional"].get("instrument_type")

            specific_gravity = data["optional"].get("specific_gravity")
            if specific_gravity is not None:
                if specific_gravity < 0:
                    raise ConversionError(
                        f"The value {specific_gravity} for SPECIFIC_GRAVITY in row {row} of sheet {sheet_name} "
                        f"is not greater than 0, as required")

            magnetic_susceptibility = data["optional"].get("magnetic_susceptibility")
            if magnetic_susceptibility is not None:
                if not str(magnetic_susceptibility).startswith("-"):
                    raise ConversionError(
                        f"The value {magnetic_susceptibility} for MAGNETIC_SUSCEPTIBILITY in row {row} of sheet {sheet_name} "
                        f"is not negative, as required")

            remark = data["optional"].get("remark")

            # make RDFLib objects of the values
            sample_iri = make_rdflib_type(data["required"]["sample_id"], "URIRef", None, Namespace(dataset_iri + "/sample/"))
            sample_material_iri = make_rdflib_type(data["required"]["sample_material"], "Concept", combined_concepts)
            sample_type_surface_iri = make_rdflib_type(data["required"]["sample_type_surface"], "Concept", combined_concepts)
            sample_mesh_size_iri = make_rdflib_type(data["required"]["sample_mesh_size"], "Concept", combined_concepts)
            soil_sample_depth_lit = make_rdflib_type(data["required"]["soil_sample_depth"], "Number")
            soil_colour_iri = make_rdflib_type(data["required"]["soil_colour"], "Concept", combined_concepts)
            soil_ph_lit = make_rdflib_type(data["required"]["soil_ph"], "Number")
            transformer = Transformer.from_crs("EPSG:32755", "EPSG:4326")
            lon, lat = transformer.transform(easting, northing)
            if elevation is not None:
                wkt = Literal(f"POINTZ({lon} {lat} {elevation})", datatype=GEO.wktLiteral)
            else:
                wkt = Literal(f"POINT({lon} {lat})", datatype=GEO.wktLiteral)
            location_survey_type_iri = get_iri_from_code(data["required"]["location_survey_type"], combined_concepts)
            collection_date_lit = make_rdflib_type(collection_date, "Date")
            dispatch_date_lit = make_rdflib_type(dispatch_date, "Date")
            if instrument_type is not None:
                instrument_type_lit = make_rdflib_type(instrument_type, "String")
            if specific_gravity is not None:
                specific_gravity_lit = make_rdflib_type(specific_gravity, "Number")
            if magnetic_susceptibility is not None:
                magnetic_susceptibility_lit = make_rdflib_type(magnetic_susceptibility, "Number")
            if remark is not None:
                remark_lit = make_rdflib_type(remark, "String")

            # make the graph
            g.add((dataset_iri, SDO.hasPart, sample_iri))

            g.add((sample_iri, RDF.type, SOSA.Sample))
            g.add((sample_iri, SDO.material, sample_material_iri))
            g.add((sample_iri, SDO.additionalType, sample_type_surface_iri))
            g.add((sample_iri, EX.meshSize, sample_mesh_size_iri))
            g.add((sample_iri, SDO.depth, soil_sample_depth_lit))
            g.add((sample_iri, SDO.color, soil_colour_iri))
            g.add((sample_iri, EX.ph, soil_colour_iri))

            ph_obs = BNode()
            ph_res = BNode()
            g.add((ph_obs, RDF.type, SOSA.Observation))
            g.add((ph_obs, SOSA.hasFeatureOfInterest, sample_iri))
            g.add((sample_iri, SOSA.isFeatureOfInterestOf, ph_obs))
            g.add((ph_obs, SOSA.observedProperty, QKINDS.PH))
            g.add((ph_obs, SOSA.hasResult, ph_res))
            g.add((ph_res, RDF.type, SOSA.Result))
            g.add((ph_res, SDO.value, soil_ph_lit))

            geom = BNode()
            g.add((sample_iri, GEO.hasGeometry, geom))  # sdo:location would be the location of the sample now
            g.add((geom, RDF.type, GEO.Geometry))
            g.add((geom, GEO.asWKT, wkt))

            g.add((sample_iri, EX.locationSurveyType, location_survey_type_iri))
            g.add((sample_iri, PROV.generatedAtTime, collection_date_lit))
            g.add((sample_iri, SDO.dateIssued, dispatch_date_lit))

            if instrument_type is not None:
                g.add((sample_iri, SOSA.madeBySensor, instrument_type_lit))

            if specific_gravity is not None:
                spec_grav_obs = BNode()
                spec_grav_res = BNode()
                g.add((spec_grav_obs, RDF.type, SOSA.Observation))
                g.add((spec_grav_obs, SOSA.hasFeatureOfInterest, sample_iri))
                g.add((sample_iri, SOSA.isFeatureOfInterestOf, spec_grav_obs))
                g.add((spec_grav_obs, SOSA.observedProperty, EX.SpecificGravity))
                g.add((spec_grav_obs, SOSA.hasResult, spec_grav_res))
                g.add((spec_grav_res, RDF.type, SOSA.Result))
                g.add((spec_grav_res, SDO.value, specific_gravity_lit))
            if magnetic_susceptibility is not None:
                mag_sup_obs = BNode()
                mag_sup_res = BNode()
                g.add((mag_sup_obs, RDF.type, SOSA.Observation))
                g.add((mag_sup_obs, SOSA.hasFeatureOfInterest, sample_iri))
                g.add((sample_iri, SOSA.isFeatureOfInterestOf, mag_sup_obs))
                g.add((mag_sup_obs, SOSA.observedProperty, QKINDS.MagneticSusceptability))
                g.add((mag_sup_obs, SOSA.hasResult, mag_sup_res))
                g.add((mag_sup_res, RDF.type, SOSA.Result))
                g.add((mag_sup_res, SDO.value, magnetic_susceptibility_lit))
            if remark is not None:
                g.add((sample_iri, RDFS.comment, remark_lit))

            sample_ids.append(data["required"]["sample_id"])

            row += 1
        else:
            break

    g.bind("ex", EX)
    g.bind("qkinds", QKINDS)

    return g, sample_ids


def extract_sheet_sample_preparation(
        wb: openpyxl.Workbook,
        laboratory_names_and_ids: Dict,
        user_sample_prep_code_ids: List[str],
        user_assay_code_ids: List[str],
        sample_ids: List[str],
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Tuple[Graph, List]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "SAMPLE_PREPARATION"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()
    job_numbers = []

    while True:
        bv = sheet[f"B{row}"].value

        if bv is not None:
            if bv == "TV19287993":
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "job_number": bv,
                        "laboratory": sheet[f"C{row}"].value,
                        "sample_prep_codes": sheet[f"D{row}"].value,  # USER_SAMPLE_PREP_CODES
                        "assay_code": sheet[f"E{row}"].value,  # USER_ASSAY_CODES
                        "sample_id": sheet[f"F{row}"].value,
                    }
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                job_number = data["required"]["job_number"]

                laboratory_name = data["required"]["laboratory"]
                if laboratory_name not in laboratory_names_and_ids.keys():
                    raise ConversionError(
                        f"The value {laboratory_name} for LABORATORY in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_LABORATORIES worksheet as required")

                sample_prep_codes = [x.strip() for x in data["required"]["sample_prep_codes"].split(";")]
                for sample_prep_code in sample_prep_codes:
                    if sample_prep_code not in user_sample_prep_code_ids:
                        raise ConversionError(
                            f"The value {sample_prep_code} for SAMPLE_PREP_CODES in row {row} of sheet {sheet_name} is "
                            f"not defined in the USER_SAMPLE_PREP_CODES worksheet as required.")

                assay_code = data["required"]["assay_code"]
                if assay_code not in user_assay_code_ids:
                    raise ConversionError(
                        f"The value {assay_code} for ASSAY_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ASSAY_CODES worksheet as required")

                sample_id = data["required"]["sample_id"]
                if sample_id not in sample_ids:
                    raise ConversionError(
                        f"The value {sample_id} for SAMPLE_ID in row {row} of sheet {sheet_name} is "
                        f"not defined in either the DRILLHOLE_SAMPLE or the SURFACE_SAMPLE worksheet as required")

                # make RDFLib objects of the values
                job_number_iri = URIRef(Namespace(dataset_iri + "/jobNumber/") + job_number)
                laboratory_iri = URIRef(Namespace(dataset_iri + "/lab/" + laboratory_names_and_ids[laboratory_name]))
                sample_prep_codes_iris = [URIRef(Namespace(dataset_iri + "/samplePrepCode/") + x) for x in sample_prep_codes]
                assay_code_iri = URIRef(Namespace(dataset_iri + "/assayCode/") + assay_code)
                sample_iri = make_rdflib_type(sample_id, "URIRef", None, Namespace(dataset_iri + "/sample/"))

                # make the graph
                g.add((dataset_iri, SDO.hasPart, job_number_iri))

                g.add((job_number_iri, RDF.type, SOSA.ObservationCollection))
                obs = BNode()
                g.add((obs, RDF.type, SOSA.Observation))
                g.add((job_number_iri, SOSA.hasMember, obs))

                qa = BNode()
                g.add((obs, PROV.qualifiedAttribution, qa))
                g.add((qa, PROV.agent, laboratory_iri))
                g.add((qa, PROV.hadRole, MININGROLES.SampleAnalyser))

                # pcs = BNode()
                # g.add((pcs, RDF.List))
                # for sample_prep_codes_iri in sample_prep_codes_iris:
                #     g.add(obs, SOSA.usedProcedure, sample_prep_codes_iri)

                pcs = BNode()
                Seq(g, pcs, sample_prep_codes_iris)
                g.add((pcs, RDFS.label, Literal("Sample Preparation codes")))
                g.add((obs, SOSA.usedProcedure, pcs))

                g.add((obs, SOSA.usedProcedure, assay_code_iri))

                g.add((obs, SOSA.hasFeatureOfInterest, sample_iri))

                g.add((dataset_iri, SDO.hasPart, sample_iri))

                job_numbers.append(bv)

                row += 1
        else:
            break

    return g, job_numbers


def extract_sheet_geochemistry_meta(
        wb: openpyxl.Workbook,
        job_numbers: List[str],
        laboratory_names_and_ids: Dict,
        user_assay_code_ids: List[str],
        analyte_ids: List[str],
        unit_of_measure_ids: List[str],
        combined_concepts: Graph,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "GEOCHEMISTRY_META"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv == "GC12345":
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "job_number": bv,
                        "laboratory": sheet[f"C{row}"].value,
                        "assay_code": sheet[f"D{row}"].value,  # USER_ASSAY_CODES
                        "analyte_code": sheet[f"E{row}"].value,
                        "unit_of_measure": sheet[f"F{row}"].value,
                        "lower_detection_limit": sheet[f"G{row}"].value,
                        "accuracy": sheet[f"H{row}"].value,
                        "preferred_result": sheet[f"J{row}"].value,
                    },
                    "optional": {
                        "upper_detection_limit": sheet[f"I{row}"].value,
                    }
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                job_number = data["required"]["job_number"]
                if job_number not in job_numbers:
                    raise ConversionError(
                        f"The value {job_number} for JOB_NUMBER in row {row} of sheet {sheet_name} is "
                        f"not present in the SAMPLE_PREPARATION job numbers worksheet as required")

                laboratory_name = data["required"]["laboratory"]
                if laboratory_name not in laboratory_names_and_ids.keys():
                    raise ConversionError(
                        f"The value {laboratory_name} for LABORATORY in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_LABORATORIES worksheet as required")

                assay_code = data["required"]["assay_code"]
                if assay_code not in user_assay_code_ids:
                    raise ConversionError(
                        f"The value {assay_code} for ASSAY_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ASSAY_CODES worksheet as required")

                analyte_code = data["required"]["analyte_code"]
                if analyte_code not in analyte_ids:
                    raise ConversionError(
                        f"The value {analyte_code} for ANALYTE_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ANALYTES worksheet as required")

                unit_of_measure = data["required"]["unit_of_measure"].split("(")[1].split(")")[0]
                if unit_of_measure not in unit_of_measure_ids:
                    raise ConversionError(
                        f"The value {unit_of_measure} for SAMPLE_ID in row {row} of sheet {sheet_name} is "
                        f"not defined in the UNITS_OF_MEASURE or the USER_UNITS_OF_MEASURE worksheets as required")

                # make RDFLib objects of the values
                job_number_iri = URIRef(Namespace(dataset_iri + "/jobNumber/") + job_number)
                laboratory_iri = URIRef(Namespace(dataset_iri + "/lab/" + laboratory_names_and_ids[laboratory_name]))
                assay_code_iri = URIRef(Namespace(dataset_iri + "/assayCode/") + assay_code)
                analyte_code_iri = URIRef(Namespace(dataset_iri + "/analyteCode/") + analyte_code)
                unit_of_measure_iri = make_rdflib_type(unit_of_measure, "Concept", combined_concepts)
                lower_detection_limit_lit = make_rdflib_type(data["required"]["lower_detection_limit"], "Number")
                accuracy_lit = make_rdflib_type(data["required"]["accuracy"], "Number")
                if data["required"].get("upper_detection_limit") is not None:
                    upper_detection_limit_lit = make_rdflib_type(data["required"].get("upper_detection_limit"), "Number")
                preferred_result_lit = make_rdflib_type(True if data["required"]["preferred_result"] == "Yes" else  False, "Boolean")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, job_number_iri))

                obs = BNode()
                g.add((obs, RDF.type, SOSA.Observation))
                g.add((job_number_iri, SOSA.hasMember, obs))

                qa = BNode()
                g.add((obs, PROV.qualifiedAttribution, qa))
                g.add((qa, PROV.agent, laboratory_iri))
                g.add((qa, PROV.hadRole, MININGROLES.SampleAnalyser))

                g.add((obs, SOSA.usedProcedure, assay_code_iri))

                g.add((obs, SOSA.observedProperty, analyte_code_iri))

                r = BNode()
                g.add((obs, SOSA.hasResult, r))
                g.add((r, RDF.type, SOSA.Result))
                g.add((r, SDO.unitCode, unit_of_measure_iri))
                g.add((r, EX.lowerDetectionLimit, lower_detection_limit_lit))
                g.add((r, SDO.marginOfError, accuracy_lit))
                if data["required"].get("upper_detection_limit") is not None:
                    g.add((r, EX.upperDetectionLimit, upper_detection_limit_lit))
                g.add((r, EX.isPreferredResult, preferred_result_lit))

                row += 1
        else:
            break

    g.bind("ex", EX)

    return g


def extract_sheet_sample_geochemistry(
        wb: openpyxl.Workbook,
        job_numbers: List[str],
        sample_ids: List[str],
        user_assay_code_ids: List[str],
        analyte_ids: List[str],
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "SAMPLE_GEOCHEMISTRY"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv == "TV19287993":
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "job_number": bv,
                        "sample_id": sheet[f"C{row}"].value,
                        "assay_code": sheet[f"D{row}"].value,  # USER_ASSAY_CODES
                        "analyte_code": sheet[f"E{row}"].value,
                        "result": sheet[f"F{row}"].value,
                    },
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                job_number = data["required"]["job_number"]
                if job_number not in job_numbers:
                    raise ConversionError(
                        f"The value {job_number} for JOB_NUMBER in row {row} of sheet {sheet_name} is "
                        f"not present in the SAMPLE_PREPARATION job numbers worksheet as required")

                sample_id = data["required"]["sample_id"]
                if sample_id not in sample_ids:
                    raise ConversionError(
                        f"The value {sample_id} for SAMPLE_ID in row {row} of sheet {sheet_name} is "
                        f"not defined in either the DRILLHOLE_SAMPLE or the SURFACE_SAMPLE worksheet as required")

                assay_code = data["required"]["assay_code"]
                if assay_code not in user_assay_code_ids:
                    raise ConversionError(
                        f"The value {assay_code} for ASSAY_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ASSAY_CODES worksheet as required")

                analyte_code = data["required"]["analyte_code"]
                if analyte_code not in analyte_ids:
                    raise ConversionError(
                        f"The value {analyte_code} for ANALYTE_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ANALYTES worksheet as required")

                result = data["required"]["result"]

                # make RDFLib objects of the values
                job_number_iri = make_rdflib_type(job_number, "URIRef", None, Namespace(dataset_iri + "/jobNumber/"))
                sample_iri = make_rdflib_type(sample_id, "URIRef", None, Namespace(dataset_iri + "/sample/"))
                assay_code_iri = make_rdflib_type(assay_code, "URIRef", None, Namespace(dataset_iri + "/assayCode/"))
                analyte_code_iri = make_rdflib_type(analyte_code, "URIRef", None, Namespace(dataset_iri + "/analyteCode/"))
                result_lit = make_rdflib_type(result, "Number")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, job_number_iri))

                obs = BNode()
                g.add((obs, RDF.type, SOSA.Observation))
                g.add((job_number_iri, SOSA.hasMember, obs))

                g.add((obs, SOSA.hasFeatureOfInterest, sample_iri))

                g.add((obs, SOSA.usedProcedure, assay_code_iri))

                g.add((obs, SOSA.observedProperty, analyte_code_iri))

                r = BNode()
                g.add((r, RDF.type, SOSA.Result))
                g.add((r, SDO.value, result_lit))

                g.add((obs, SOSA.hasResult, r))

                row += 1
        else:
            break

    g.bind("ex", EX)

    return g


def extract_sheet_qaqc_meta(
        wb: openpyxl.Workbook,
        job_numbers: List[str],
        laboratory_names_and_ids: Dict,
        user_assay_code_ids: List[str],
        analyte_ids: List[str],
        unit_of_measure_ids: List[str],
        combined_concepts: Graph,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "QAQC_META"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv == "GC12345":
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "job_number": bv,
                        "laboratory": sheet[f"C{row}"].value,
                        "assay_code": sheet[f"D{row}"].value,  # USER_ASSAY_CODES
                        "analyte_code": sheet[f"E{row}"].value,
                        "unit_of_measure": sheet[f"F{row}"].value,
                        "lower_detection_limit": sheet[f"G{row}"].value,
                        "accuracy": sheet[f"H{row}"].value,
                        "preferred_result": sheet[f"J{row}"].value,
                    },
                    "optional": {
                        "upper_detection_limit": sheet[f"I{row}"].value,
                    }
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                job_number = data["required"]["job_number"]
                if job_number not in job_numbers:
                    raise ConversionError(
                        f"The value {job_number} for JOB_NUMBER in row {row} of sheet {sheet_name} is "
                        f"not present in the SAMPLE_PREPARATION job numbers worksheet as required")

                laboratory_name = data["required"]["laboratory"]
                if laboratory_name not in laboratory_names_and_ids.keys():
                    raise ConversionError(
                        f"The value {laboratory_name} for LABORATORY in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_LABORATORIES worksheet as required")

                assay_code = data["required"]["assay_code"]
                if assay_code not in user_assay_code_ids:
                    raise ConversionError(
                        f"The value {assay_code} for ASSAY_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ASSAY_CODES worksheet as required")

                analyte_code = data["required"]["analyte_code"]
                if analyte_code not in analyte_ids:
                    raise ConversionError(
                        f"The value {analyte_code} for ANALYTE_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ANALYTES worksheet as required")

                unit_of_measure = data["required"]["unit_of_measure"].split("(")[1].split(")")[0]
                if unit_of_measure not in unit_of_measure_ids:
                    raise ConversionError(
                        f"The value {unit_of_measure} for SAMPLE_ID in row {row} of sheet {sheet_name} is "
                        f"not defined in the UNITS_OF_MEASURE or the USER_UNITS_OF_MEASURE worksheets as required")

                # make RDFLib objects of the values
                job_number_iri = URIRef(Namespace(dataset_iri + "/jobNumber/") + job_number)
                laboratory_iri = URIRef(Namespace(dataset_iri + "/lab/" + laboratory_names_and_ids[laboratory_name]))
                assay_code_iri = URIRef(Namespace(dataset_iri + "/assayCode/") + assay_code)
                analyte_code_iri = URIRef(Namespace(dataset_iri + "/analyteCode/") + analyte_code)
                unit_of_measure_iri = make_rdflib_type(unit_of_measure, "Concept", combined_concepts)
                lower_detection_limit_lit = make_rdflib_type(data["required"]["lower_detection_limit"], "Number")
                accuracy_lit = make_rdflib_type(data["required"]["accuracy"], "Number")
                if data["required"].get("upper_detection_limit") is not None:
                    upper_detection_limit_lit = make_rdflib_type(data["required"].get("upper_detection_limit"), "Number")
                preferred_result_lit = make_rdflib_type(True if data["required"]["preferred_result"] == "Yes" else  False, "Boolean")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, job_number_iri))

                obs = BNode()
                g.add((obs, RDF.type, SOSA.Observation))
                g.add((job_number_iri, SOSA.hasMember, obs))

                qa = BNode()
                g.add((obs, PROV.qualifiedAttribution, qa))
                g.add((qa, PROV.agent, laboratory_iri))
                g.add((qa, PROV.hadRole, MININGROLES.SampleAnalyser))

                g.add((obs, SOSA.usedProcedure, assay_code_iri))

                g.add((obs, SOSA.observedProperty, analyte_code_iri))

                r = BNode()
                g.add((obs, SOSA.hasResult, r))
                g.add((r, RDF.type, SOSA.Result))
                g.add((r, SDO.unitCode, unit_of_measure_iri))
                g.add((r, EX.lowerDetectionLimit, lower_detection_limit_lit))
                g.add((r, SDO.marginOfError, accuracy_lit))
                if data["required"].get("upper_detection_limit") is not None:
                    g.add((r, EX.upperDetectionLimit, upper_detection_limit_lit))
                g.add((r, EX.isPreferredResult, preferred_result_lit))

                row += 1
        else:
            break

    g.bind("ex", EX)

    return g


def extract_sheet_qaqc_geochemistry(
        wb: openpyxl.Workbook,
        job_numbers: List[str],
        sample_ids: List[str],
        user_assay_code_ids: List[str],
        analyte_ids: List[str],
        combined_concepts: Graph,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "QAQC_GEOCHEMISTRY"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv == "GC12345":
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "job_number": bv,
                        "sample_id": sheet[f"C{row}"].value,
                        "assay_code": sheet[f"D{row}"].value,  # USER_ASSAY_CODES
                        "analyte_code": sheet[f"E{row}"].value,
                        "result": sheet[f"F{row}"].value,
                        "orig_sample_id": sheet[f"G{row}"].value,
                        "qaqc_type": sheet[f"H{row}"].value,
                        "standard_id": sheet[f"I{row}"].value,
                        "standard_provider": sheet[f"J{row}"].value,
                    },
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                job_number = data["required"]["job_number"]
                if job_number not in job_numbers:
                    raise ConversionError(
                        f"The value {job_number} for JOB_NUMBER in row {row} of sheet {sheet_name} is "
                        f"not present in the SAMPLE_PREPARATION job numbers worksheet as required")

                sample_id = data["required"]["sample_id"]
                if sample_id not in sample_ids:
                    raise ConversionError(
                        f"The value {sample_id} for SAMPLE_ID in row {row} of sheet {sheet_name} is "
                        f"not defined in either the DRILLHOLE_SAMPLE or the SURFACE_SAMPLE worksheet as required")

                assay_code = data["required"]["assay_code"]
                if assay_code not in user_assay_code_ids:
                    raise ConversionError(
                        f"The value {assay_code} for ASSAY_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ASSAY_CODES worksheet as required")

                analyte_code = data["required"]["analyte_code"]
                if analyte_code not in analyte_ids:
                    raise ConversionError(
                        f"The value {analyte_code} for ANALYTE_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ANALYTES worksheet as required")

                result = data["required"]["result"]

                orig_sample_id = data["required"]["orig_sample_id"]

                validate_code(
                    data["required"]["qaqc_type"], "QAQC", "QAQC", row,
                    sheet_name,
                    combined_concepts
                )
                qaqc_type = data["required"]["qaqc_type"]

                standard_id = data["required"]["standard_id"]

                standard_provider = data["required"]["standard_provider"]

                # make RDFLib objects of the values
                job_number_iri = make_rdflib_type(job_number, "URIRef", None, Namespace(dataset_iri + "/jobNumber/"))
                sample_iri = make_rdflib_type(sample_id, "URIRef", None, Namespace(dataset_iri + "/sample/"))
                assay_code_iri = make_rdflib_type(assay_code, "URIRef", None, Namespace(dataset_iri + "/assayCode/"))
                analyte_code_iri = make_rdflib_type(analyte_code, "URIRef", None, Namespace(dataset_iri + "/analyteCode/"))
                result_lit = make_rdflib_type(result, "Number")
                orig_sample_iri = make_rdflib_type(orig_sample_id, "URIRef", None, Namespace(dataset_iri + "/sample/"))
                qaqc_type_iri = make_rdflib_type(qaqc_type, "Concept", combined_concepts)
                standard_id_lit = make_rdflib_type(standard_id, "String")
                standard_provider_lit = make_rdflib_type(standard_provider, "String")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, job_number_iri))

                obs = BNode()
                g.add((obs, RDF.type, SOSA.Observation))
                g.add((job_number_iri, SOSA.hasMember, obs))

                g.add((obs, SOSA.hasFeatureOfInterest, sample_iri))

                g.add((obs, SOSA.usedProcedure, assay_code_iri))

                g.add((obs, SOSA.observedProperty, analyte_code_iri))

                r = BNode()
                g.add((r, RDF.type, SOSA.Result))
                g.add((r, SDO.value, result_lit))

                g.add((obs, SOSA.hasResult, r))

                g.add((sample_iri, SOSA.isSampleOf, orig_sample_iri))
                g.add((sample_iri, SDO.additionalType, qaqc_type_iri))
                p = BNode()
                g.add((p, RDF.type, SOSA.Procedure))
                g.add((p, SDO.identifier, standard_id_lit))
                g.add((p, SDO.author, standard_provider_lit))
                g.add((obs, SOSA.usedProcedure, p))

                row += 1
        else:
            break

    g.bind("ex", EX)

    return g


def extract_sheet_sample_pxrf(
        wb: openpyxl.Workbook,
        sample_ids: List[str],
        analyte_ids: List[str],
        unit_of_measure_ids: List[str],
        combined_concepts: Graph,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "SAMPLE_PXRF"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()
    obs_col_bn = BNode()
    g.add((obs_col_bn, RDF.type, SOSA.ObservationCollection))
    g.add((dataset_iri, SDO.hasPart, obs_col_bn))

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv == "SS12345":
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "sample_id": bv,
                        "reading_no": sheet[f"C{row}"].value,
                        "filter_beam_setting": sheet[f"E{row}"].value,
                        "xrf_beam1_time": sheet[f"F{row}"].value,
                        "xrf_beam2_time": sheet[f"G{row}"].value,
                        "xrf_beam3_time": sheet[f"H{row}"].value,
                        "xrf_elapsed_time": sheet[f"I{row}"].value,
                        "xrf_instrument_type": sheet[f"J{row}"].value,
                        "analyte_code": sheet[f"K{row}"].value,
                        "unit_of_measure": sheet[f"L{row}"].value,
                        "result": sheet[f"M{row}"].value,
                    },
                    "optional": {
                        "mode": sheet[f"D{row}"].value,
                    }
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                sample_id = data["required"]["sample_id"]
                if sample_id not in sample_ids:
                    raise ConversionError(
                        f"The value {sample_id} for SAMPLE_ID in row {row} of sheet {sheet_name} is "
                        f"not defined in either the DRILLHOLE_SAMPLE or the SURFACE_SAMPLE worksheet as required")

                if data["optional"].get("mode") is not None:
                    result = data["optional"]["mode"]

                filter_beam_setting = data["required"]["filter_beam_setting"]
                xrf_beam1_time = data["required"]["xrf_beam1_time"]
                xrf_beam2_time = data["required"]["xrf_beam2_time"]
                xrf_beam3_time = data["required"]["xrf_beam3_time"]
                xrf_elapsed_time = data["required"]["xrf_elapsed_time"]
                xrf_instrument_type = data["required"]["xrf_instrument_type"]

                analyte_code = data["required"]["analyte_code"]
                if analyte_code not in analyte_ids:
                    raise ConversionError(
                        f"The value {analyte_code} for ANALYTE_CODE in row {row} of sheet {sheet_name} is "
                        f"not defined in the USER_ANALYTES worksheet as required")

                unit_of_measure = data["required"]["unit_of_measure"].split("(")[1].split(")")[0]
                if unit_of_measure not in unit_of_measure_ids:
                    raise ConversionError(
                        f"The value {unit_of_measure} for SAMPLE_ID in row {row} of sheet {sheet_name} is "
                        f"not defined in the UNITS_OF_MEASURE or the USER_UNITS_OF_MEASURE worksheets as required")

                result = data["required"]["result"]

                # make RDFLib objects of the values
                sample_iri = make_rdflib_type(sample_id, "URIRef", None, Namespace(dataset_iri + "/sample/"))
                json_val = {
                    "FILTER_BEAM_SETTING": filter_beam_setting,
                    "XRF_BEAM1_TIME": xrf_beam1_time,
                    "XRF_BEAM2_TIME": xrf_beam2_time,
                    "XRF_BEAM3_TIME": xrf_beam3_time,
                    "XRF_ELAPSED_TIME": xrf_elapsed_time,
                    "XRF_INSTRUMENT_TYPE": xrf_instrument_type
                }
                procedure_lit = Literal(json.dumps(json_val), datatype=RDF.JSON)
                analyte_code_iri = make_rdflib_type(analyte_code, "URIRef", None, Namespace(dataset_iri + "/analyteCode/"))
                uom_iri = make_rdflib_type(unit_of_measure, "Concept", combined_concepts)
                result_lit = make_rdflib_type(result, "Number")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, sample_iri))

                obs = BNode()
                g.add((obs, RDF.type, SOSA.Observation))
                g.add((obs, SOSA.hasFeatureOfInterest, sample_iri))
                g.add((obs_col_bn, SOSA.hasMember, obs))

                procedure_bn = BNode()
                g.add((procedure_bn, RDF.type, SOSA.Procedure))
                g.add((procedure_bn, SDO.name, Literal("XRF Analysis")))
                g.add((procedure_bn, SDO.description, procedure_lit))
                g.add((obs, SOSA.usedProcedure, procedure_bn))

                g.add((obs, SOSA.observedProperty, analyte_code_iri))

                r = BNode()
                g.add((r, RDF.type, SOSA.Result))
                g.add((r, SDO.unitCode, uom_iri))
                g.add((r, SDO.value, result_lit))
                g.add((obs, SOSA.hasResult, r))

                row += 1
        else:
            break

    g.bind("ex", EX)

    return g


def extract_sheet_lith_dictionary(
        wb: openpyxl.Workbook,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Tuple[Graph, List]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "LITH_DICTIONARY"
    sheet = wb[sheet_name]

    row = 9

    lith_codes = []
    g = Graph(bind_namespaces="rdflib")

    lith_cs_iri = URIRef(str(dataset_iri) + "/lithology-cs")
    g.add((lith_cs_iri, RDF.type, SKOS.ConceptScheme))
    g.add((lith_cs_iri, SKOS.prefLabel, Literal(f"Dataset {dataset_iri} lithology vocabulary")))

    while True:
        v = sheet[f"B{row}"].value
        if v is not None:
            their_label = sheet[f"A{row}"].value
            if their_label is None:
                raise ConversionError(
                    f"The value for COMP_LITH on row {row} of the worksheet LITH_DICTIONARY must not be null")
            their_lith_code = v
            gsq_label = sheet[f"D{row}"].value
            if gsq_label is None:
                raise ConversionError(
                    f"The value for GSQ_LITH_MATCH on row {row} of the worksheet LITH_DICTIONARY must not be null")
            gsq_lith_code = sheet[f"E{row}"].value
            if gsq_lith_code is None:
                raise ConversionError(
                    f"The value for GSQ_CODE_MATCH on row {row} of the worksheet LITH_DICTIONARY must not be null")

            lith_codes.append(v)

            # make RDF types
            lith_iri = make_rdflib_type(gsq_lith_code, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
            lith_pl_lit = make_rdflib_type(gsq_label, "String")
            lith_al_lit = make_rdflib_type(their_label, "String")

            # make graph
            g.add((lith_iri, RDF.type, SKOS.Concept))
            g.add((lith_iri, SKOS.prefLabel, lith_pl_lit))
            if their_lith_code != gsq_lith_code:
                g.add((lith_iri, SKOS.notation, Literal(their_lith_code, datatype=dataset_iri)))
            if their_label != gsq_label:
                g.add((lith_iri, SKOS.altLabel, lith_al_lit))
            g.add((lith_iri, SKOS.topConceptOf, lith_cs_iri))
            g.add((lith_cs_iri, SKOS.hasTopConcept, lith_iri))

            row += 1
        else:
            break

    return g, lith_codes


def extract_sheet_min_dictionary(
        wb: openpyxl.Workbook,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Tuple[Graph, List]:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "MIN_DICTIONARY"
    sheet = wb[sheet_name]

    row = 9

    min_codes = []
    g = Graph(bind_namespaces="rdflib")

    min_cs_iri = URIRef(str(dataset_iri) + "/minerals-cs")
    g.add((min_cs_iri, RDF.type, SKOS.ConceptScheme))
    g.add((min_cs_iri, SKOS.prefLabel, Literal(f"Dataset {dataset_iri} minerals vocabulary")))

    while True:
        v = sheet[f"B{row}"].value
        if v is not None:
            their_label = sheet[f"A{row}"].value
            if their_label is None:
                raise ConversionError(
                    f"The value for COMP_MIN on row {row} of the worksheet MIN_DICTIONARY must not be null")
            their_lith_code = v
            gsq_label = sheet[f"C{row}"].value
            if gsq_label is None:
                raise ConversionError(
                    f"The value for GSQ_MIN_MATCH on row {row} of the worksheet MIN_DICTIONARY must not be null")
            gsq_lith_code = sheet[f"D{row}"].value
            if gsq_lith_code is None:
                raise ConversionError(
                    f"The value for GSQ_CODE_MATCH on row {row} of the worksheet MIN_DICTIONARY must not be null")

            min_codes.append(v)

            # make RDF types
            min_iri = make_rdflib_type(gsq_lith_code, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
            min_pl_lit = make_rdflib_type(gsq_label, "String")
            min_al_lit = make_rdflib_type(their_label, "String")

            # make graph
            g.add((min_iri, RDF.type, SKOS.Concept))
            g.add((min_iri, SKOS.prefLabel, min_pl_lit))
            if their_lith_code != gsq_lith_code:
                g.add((min_iri, SKOS.notation, Literal(their_lith_code, datatype=dataset_iri)))
            if their_label != gsq_label:
                g.add((min_iri, SKOS.altLabel, min_al_lit))
            g.add((min_iri, SKOS.topConceptOf, min_cs_iri))
            g.add((min_cs_iri, SKOS.hasTopConcept, min_iri))

            row += 1
        else:
            break

    return g, min_codes


def extract_sheet_drillhole_lithology(
        wb: openpyxl.Workbook,
        drillhole_ids: List[str],
        lith_code_ids: List[str],
        min_code_ids: List[str],
        combined_concepts: Graph,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "DRILLHOLE_LITHOLOGY"
    sheet = wb[sheet_name]

    row = 9
    g = Graph()

    while True:
        bv = sheet[f"B{row}"].value
        if bv is not None:
            if bv == "DD12345":
                row += 1
                continue
            else:
                # make vars of all the sheet values
                data = {
                    "required": {
                        "drillhole_id": bv,
                        "from": sheet[f"C{row}"].value,
                        "to": sheet[f"D{row}"].value,

                        "rock_1": sheet[f"I{row}"].value,
                        "rock_2": sheet[f"K{row}"].value,

                        "alt_type": sheet[f"S{row}"].value,
                        "alt_intensity": sheet[f"T{row}"].value,
                    },
                    "optional": {
                        "recovered_amount": sheet[f"E{row}"].value,
                        "weathering": sheet[f"F{row}"].value,
                        "colour": sheet[f"G{row}"].value,
                        "colour_shade": sheet[f"H{row}"].value,

                        "rock_1_abund": sheet[f"J{row}"].value,
                        "rock_2_abund": sheet[f"L{row}"].value,

                        "min_1": sheet[f"M{row}"].value,
                        "min_1_abund": sheet[f"N{row}"].value,
                        "min_2": sheet[f"O{row}"].value,
                        "min_2_abund": sheet[f"P{row}"].value,
                        "min_3": sheet[f"Q{row}"].value,
                        "min_3_abund": sheet[f"R{row}"].value,

                        "alt_min_1": sheet[f"U{row}"].value,
                        "alt_min_1_abund": sheet[f"V{row}"].value,
                        "alt_min_2": sheet[f"W{row}"].value,
                        "alt_min_abund_2": sheet[f"X{row}"].value,

                        "vein_composition": sheet[f"Y{row}"].value,
                        "vein_description": sheet[f"Z{row}"].value,
                        "vein_percent": sheet[f"AA{row}"].value,
                        "structure": sheet[f"AB{row}"].value,
                        "texture": sheet[f"AC{row}"].value,
                        "grain_size": sheet[f"AD{row}"].value,
                        "remark": sheet[f"AE{row}"].value,
                    }
                }

                # check required sheet values are present
                for k, v in data["required"].items():
                    if v is None:
                        raise ConversionError(
                            f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

                # value validation
                # cross-sheet validation
                drillhole_id = str(data["required"]["drillhole_id"])
                if drillhole_id not in drillhole_ids:
                    raise ConversionError(
                        f"The value {drillhole_id} for DRILLHOLE_ID in row {row} of sheet {sheet_name} "
                        f"is not present on sheet DRILLHOLE_LOCATION in the DRILLHOLE_ID column, as required")

                depth_from = data["required"]["from"]
                if depth_from < 0:
                    raise ConversionError(
                        f"The value {depth_from} for FROM in row {row} of sheet {sheet_name} is not greater or equal to "
                        f"zero as required")

                depth_to = data["required"]["to"]
                if depth_to <= depth_from:
                    raise ConversionError(
                        f"The value {depth_to} for TO in row {row} of sheet {sheet_name} is not greater or equal to "
                        f"the FROM value as required")

                rock_1 = data["required"]["rock_1"]
                if rock_1 not in lith_code_ids:
                    raise ConversionError(
                        f"The value {rock_1} for ROCK_1 in row {row} of sheet {sheet_name} defined"
                        f"in the worksheet LITH_DICTIONARY in column B")

                rock_2 = data["required"]["rock_2"]
                if rock_2 not in lith_code_ids:
                    raise ConversionError(
                        f"The value {rock_2} for ROCK_2 in row {row} of sheet {sheet_name} defined"
                        f"in the worksheet LITH_DICTIONARY in column B")

                validate_code(
                    data["required"]["alt_type"],
                    "ALTERATION",
                    "ALT_TYPE",
                    row,
                    sheet_name,
                    combined_concepts
                )
                alt_type = data["required"]["alt_type"]

                alt_intensity = data["required"]["alt_intensity"]

                recovered_amount = data["optional"].get("recovered_amount")
                if recovered_amount is not None:
                    try:
                        recovered_amount = float(recovered_amount)
                    except ValueError:
                        raise ConversionError(
                            f"The value {recovered_amount} for RECOVERED_AMOUNT in row {row} of sheet {sheet_name} "
                            f"cannot be converted to a number")

                    if recovered_amount < 0 or recovered_amount > 100:
                        raise ConversionError(
                            f"The value {rock_2} for RECOVERED_AMOUNT in row {row} of sheet {sheet_name} "
                            f"is not between 0 and 100, as required")

                weathering = data["optional"].get("weathering")
                if weathering is not None:
                    validate_code(
                        data["optional"]["weathering"],
                        "WEATHERING",
                        "WEATHERING",
                        row,
                        sheet_name,
                        combined_concepts
                    )

                colour = data["optional"].get("colour")
                if colour is not None:
                    validate_code(
                        colour,
                        "COLOUR",
                        "COLOUR",
                        row,
                        sheet_name,
                        combined_concepts
                    )

                colour_shade = data["optional"].get("colour_shade")

                rock_1_abund = data["optional"].get("rock_1_abund")
                if rock_1_abund is not None:
                    try:
                        rock_1_abund = float(rock_1_abund)
                    except ValueError:
                        raise ConversionError(
                            f"The value {rock_1_abund} for ROCK_1_ABUND in row {row} of sheet {sheet_name} "
                            f"cannot be converted to a number")
                    if not 0 < rock_1_abund < 100:
                        raise ConversionError(
                            f"The value {rock_1_abund} for ROCK_1_ABUND in row {row} of sheet {sheet_name} "
                            f"is a percentage and must be between 0 and 100")

                rock_2_abund = data["optional"].get("rock_2_abund")
                if rock_2_abund is not None:
                    try:
                        rock_2_abund = float(rock_2_abund)
                    except ValueError:
                        raise ConversionError(
                            f"The value {rock_2_abund} for ROCK_2_ABUND in row {row} of sheet {sheet_name} "
                            f"cannot be converted to a number")
                    if not 0 < rock_2_abund < 100:
                        raise ConversionError(
                            f"The value {rock_2_abund} for ROCK_2_ABUND in row {row} of sheet {sheet_name} "
                            f"is a percentage and must be between 0 and 100")

                min_1 = data["optional"].get("min_1")
                if min_1 is not None:
                    if min_1 not in min_code_ids:
                        raise ConversionError(
                            f"The value {min_1} for MIN_1 in row {row} of sheet {sheet_name}, if given "
                            f"must be defined in the worksheet MIN_DICTIONARY in column B")

                min_1_abund = data["optional"].get("min_1_abund")
                if min_1_abund is not None:
                    try:
                        min_1_abund = float(min_1_abund)
                    except ValueError:
                        raise ConversionError(
                            f"The value {min_1_abund} for PRIM_MIN_ABUND_1 in row {row} of sheet {sheet_name}, "
                            f"if given must be convertable into a number")
                    if not 0 < min_1_abund < 100:
                        raise ConversionError(
                            f"The value {min_1_abund} for PRIM_MIN_ABUND_1 in row {row} of sheet {sheet_name}, "
                            f"if given, must be a percentage between 0 and 100")

                min_2 = data["optional"].get("min_2")
                if min_2 is not None:
                    if min_2 not in min_code_ids:
                        raise ConversionError(
                            f"The value {min_2} for MIN_2 in row {row} of sheet {sheet_name}, if given "
                            f"must be defined in the worksheet MIN_DICTIONARY in column B")

                min_2_abund = data["optional"].get("min_2_abund")
                if min_2_abund is not None:
                    try:
                        min_2_abund = float(min_2_abund)
                    except ValueError:
                        raise ConversionError(
                            f"The value {min_2_abund} for PRIM_MIN_ABUND_2 in row {row} of sheet {sheet_name}, "
                            f"if given must be convertable into a number")
                    if not 0 < min_2_abund < 100:
                        raise ConversionError(
                            f"The value {min_2_abund} for PRIM_MIN_ABUND_2 in row {row} of sheet {sheet_name}, "
                            f"if given, must be a percentage between 0 and 100")

                min_3 = data["optional"].get("min_3")
                if min_3 is not None:
                    if min_3 not in min_code_ids:
                        raise ConversionError(
                            f"The value {min_3} for MIN_3 in row {row} of sheet {sheet_name}, if given "
                            f"must be defined in the worksheet MIN_DICTIONARY in column B")

                min_3_abund = data["optional"].get("min_3_abund")
                if min_3_abund is not None:
                    try:
                        min_3_abund = float(min_3_abund)
                    except ValueError:
                        raise ConversionError(
                            f"The value {min_3_abund} for PRIM_MIN_ABUND_3 in row {row} of sheet {sheet_name}, "
                            f"if given must be convertable into a number")
                    if not 0 < min_3_abund < 100:
                        raise ConversionError(
                            f"The value {min_3} for PRIM_MIN_ABUND_3 in row {row} of sheet {sheet_name}, "
                            f"if given, must be a percentage between 0 and 100")

                alt_min_1 = data["optional"].get("alt_min_1")
                if min_1 is not None:
                    if alt_min_1 not in min_code_ids:
                        raise ConversionError(
                            f"The value {alt_min_1} for ATL_MIN_1 in row {row} of sheet {sheet_name}, if given "
                            f"must be defined in the worksheet MIN_DICTIONARY in column B")

                alt_min_1_abund = data["optional"].get("alt_min_1_abund")
                if alt_min_1_abund is not None:
                    try:
                        alt_min_1_abund = float(alt_min_1_abund)
                    except ValueError:
                        raise ConversionError(
                            f"The value {alt_min_1_abund} for ATL_MIN_ABUND_1 in row {row} of sheet {sheet_name}, "
                            f"if given must be convertable into a number")
                    if not 0 < alt_min_1_abund < 100:
                        raise ConversionError(
                            f"The value {alt_min_1_abund} for ATL_MIN_ABUND_1 in row {row} of sheet {sheet_name}, "
                            f"if given, must be a percentage between 0 and 100")

                alt_min_2 = data["optional"].get("alt_min_2")
                if alt_min_2 is not None:
                    if alt_min_2 not in min_code_ids:
                        raise ConversionError(
                            f"The value {alt_min_2} for ATL_MIN_2 in row {row} of sheet {sheet_name}, if given "
                            f"must be defined in the worksheet MIN_DICTIONARY in column B")

                alt_min_2_abund = data["optional"].get("alt_min_2_abund")
                if alt_min_2_abund is not None:
                    try:
                        alt_min_abund_2 = float(alt_min_2_abund)
                    except ValueError:
                        raise ConversionError(
                            f"The value {alt_min_2_abund} for ATL_MIN_ABUND_1 in row {row} of sheet {sheet_name}, "
                            f"if given must be convertable into a number")
                    if not 0 < alt_min_abund_2 < 100:
                        raise ConversionError(
                            f"The value {alt_min_abund_2} for ATL_MIN_ABUND_1 in row {row} of sheet {sheet_name}, "
                            f"if given, must be a percentage between 0 and 100")

                vein_composition = data["optional"].get("vein_composition")
                if vein_composition is not None:
                    if vein_composition not in min_code_ids:
                        raise ConversionError(
                            f"The value {vein_composition} for ATL_MIN_1 in row {row} of sheet {sheet_name}, if given "
                            f"must be defined in the worksheet MIN_DICTIONARY in column B")

                vein_description = data["optional"].get("vein_description")

                vein_percent = data["optional"].get("vein_percent")
                if vein_percent is not None:
                    try:
                        vein_percent = float(vein_percent)
                    except ValueError:
                        raise ConversionError(
                            f"The value {vein_percent} for VEIN_PERCENT in row {row} of sheet {sheet_name}, "
                            f"if given must be convertable into a number")
                    if not 0 < vein_percent < 100:
                        raise ConversionError(
                            f"The value {vein_percent} for VEIN_PERCENT in row {row} of sheet {sheet_name}, "
                            f"if given, must be a percentage between 0 and 100")

                structure = data["optional"].get("structure")
                if structure is not None:
                    validate_code(
                        structure,
                        "STRUCTURAL_FEATURE",
                        "STRUCTURE",
                        row,
                        sheet_name,
                        combined_concepts
                    )

                texture = data["optional"].get("texture")
                if texture is not None:
                    validate_code(
                        texture,
                        "TEXTURE",
                        "TEXTURE",
                        row,
                        sheet_name,
                        combined_concepts
                    )

                grain_size = data["optional"].get("grain_size")
                if grain_size is not None:
                    validate_code(
                        grain_size,
                        "GRAIN_SIZE",
                        "GRAIN_SIZE",
                        row,
                        sheet_name,
                        combined_concepts
                    )

                remark = data["optional"].get("remark")

                # make RDFLib objects of the values
                drillhole_iri = URIRef(QLDBORES + drillhole_id)
                depth_from_lit = make_rdflib_type(depth_from, "Number")
                depth_to_lit = make_rdflib_type(depth_to, "Number")

                rock_1_iri = make_rdflib_type(rock_1, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
                rock_1_abund_lit = make_rdflib_type(rock_1_abund, "Number")
                rock_2_iri = make_rdflib_type(rock_2, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
                rock_2_abund_lit = make_rdflib_type(rock_2_abund, "Number")

                min_1_iri = make_rdflib_type(min_1, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
                min_1_abund_lit = make_rdflib_type(min_1_abund, "Number")
                min_2_iri = make_rdflib_type(min_2, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
                min_2_abund_lit = make_rdflib_type(min_2_abund, "Number")
                min_3_iri = make_rdflib_type(min_3, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
                min_3_abund_lit = make_rdflib_type(min_3_abund, "Number")

                alt_type_iri = make_rdflib_type(alt_type, "Concept", combined_concepts)
                alt_intensity_lit = make_rdflib_type(alt_intensity, "String")

                alt_min_1_iri = make_rdflib_type(alt_min_1, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
                alt_min_1_abund_lit = make_rdflib_type(alt_min_1_abund, "Number")
                alt_min_2_iri = make_rdflib_type(alt_min_2, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
                alt_min_2_abund_lit = make_rdflib_type(alt_min_2_abund, "Number")

                recovered_amount_lit = make_rdflib_type(recovered_amount, "Number")
                weathering_iri = make_rdflib_type(weathering, "Concept", combined_concepts)
                colour_iri = make_rdflib_type(colour, "Concept", combined_concepts)
                colour_shade_lit = make_rdflib_type(colour_shade, "Number")

                vein_composition_iri = make_rdflib_type(vein_composition, "URIRef", None, Namespace(dataset_iri + "/lithology/"))
                vein_description_lit = make_rdflib_type(vein_description, "String")
                vein_percent_lit = make_rdflib_type(vein_percent, "Number")

                structure_iri = make_rdflib_type(structure, "Concept", combined_concepts)
                texture_iri = make_rdflib_type(texture, "Concept", combined_concepts)
                grain_size_iri = make_rdflib_type(grain_size, "Concept", combined_concepts)

                remark_lit = make_rdflib_type(remark, "String")

                # make the graph
                g.add((dataset_iri, SDO.hasPart, drillhole_iri))

                bh = BNode()
                g.add((bh, RDF.type, BORE.BoreholeInterval))
                g.add((drillhole_iri, SDO.hasPart, bh))

                bi = BNode()
                g.add((bi, RDF.type, BORE.BoreholeInterval))
                g.add((bi, SDO.depth, depth_from_lit))
                g.add((bi, SDO.depth, depth_to_lit))
                g.add((bh, SDO.hasPart, bi))

                s = BNode()
                g.add((s, RDF.type, SOSA.Sample))
                g.add((s, SOSA.isSampleOf, bi))

                oc = BNode()
                g.add((oc, RDF.type, SOSA.ObservationCollection))

                length = URIRef("http://qudt.org/vocab/quantitykind/Length")
                m = URIRef("http://qudt.org/vocab/unit/M")
                pc = URIRef("http://qudt.org/vocab/unit/PERCENT")
                alteration = URIRef("https://linked.data.gov.au/def/observable-properties/geological-unit-alteration")

                material_observations = [
                    # name, op, value, unit, desc
                    (Literal("Recovered Amount"), length, recovered_amount_lit, m, None),
                    (Literal("Weathering"), GEOSAMPLE.weathering, weathering_iri, None, None),
                    (Literal("Colour"), GEOSAMPLE.colour, colour_iri, None, colour_shade_lit),
                    (Literal("Rock 1"), rock_1_iri, rock_1_abund_lit, pc, None),
                    (Literal("Rock 2"), rock_2_iri, rock_2_abund_lit, pc, None),
                    (Literal("Mineral 1"), min_1_iri, min_1_abund_lit, pc, None),
                    (Literal("Mineral 2"), min_2_iri, min_2_abund_lit, pc, None),
                    (Literal("Mineral 3"), min_3_iri, min_3_abund_lit, pc, None),
                    (Literal("Alteration Type"), alteration, alt_type_iri, None, alt_intensity_lit),
                    (Literal("Alteration Mineral 1"), alt_min_1_iri, alt_min_1_abund_lit, pc, None),
                    (Literal("Alteration Mineral 2"), alt_min_2_iri, alt_min_2_abund_lit, pc, None),
                    (Literal("Vein Composition"), vein_composition_iri, vein_percent_lit, pc, vein_description_lit),
                    (Literal("Structure"), GEOSAMPLE.structure, structure_iri, None, vein_description_lit),
                    (Literal("Texture"), GEOSAMPLE.texture, texture_iri, None, vein_description_lit),
                    (Literal("Grain Size"), GEOSAMPLE.grainSize, grain_size_iri, None, vein_description_lit),
                ]

                for n, op, v, u, d in material_observations:
                    g += make_observation(oc, n, op, v, u, d)

                if remark_lit is not None:
                    g.add((oc, RDFS.comment, remark_lit))

                row += 1
        else:
            break

    g.bind("ex", EX)

    return g


def extract_sheet_drillhole_structure(
        wb: openpyxl.Workbook,
        drillhole_ids: List[str],
        combined_concepts: Graph,
        dataset_iri: URIRef,
        template_version: Optional[str] = None
) -> Graph:
    if template_version is None:
        template_version = check_template_version_supported(wb)

    sheet_name = "DRILLHOLE_STRUCTURE"
    sheet = wb[sheet_name]

    row = 9
    if sheet["B9"].value == "DD12345":
        row = 10

    g = Graph()

    while True:
        if sheet[f"B{row}"].value is not None:
            # make vars of all the sheet values
            data = {
                "required": {
                    "drillhole_id": sheet[f"B{row}"].value,
                    "measurement_depth": float(sheet[f"C{row}"].value),
                    "structure": sheet[f"D{row}"].value,
                    "dip": int(sheet[f"E{row}"].value),
                    "dip_direction": int(sheet[f"F{row}"].value),
                    "alpha_angle": int(sheet[f"G{row}"].value),
                    "beta_angle": int(sheet[f"H{row}"].value),
                    "azimuth": int(sheet[f"I{row}"].value),
                },
                "optional": {
                    "remark": sheet[f"J{row}"].value,
                }
            }

            # check required sheet values are present
            for k, v in data["required"].items():
                if v is None:
                    raise ConversionError(
                        f"For each row in the {sheet_name} worksheet, you must supply a {k.upper()} value")

            # value validation
            drillhole_id = str(data["required"]["drillhole_id"])
            if drillhole_id not in drillhole_ids:
                raise ConversionError(
                    f"The value {drillhole_id} for DRILLHOLE_ID in row {row} of sheet {sheet_name} "
                    f"is not present on sheet DRILLHOLE_LOCATION in the DRILLHOLE_ID column, as required")

            measurement_depth = data["required"]["measurement_depth"]
            if type(measurement_depth) not in [float, int] and measurement_depth < 0:
                raise ConversionError(
                    f"The value {measurement_depth} for MEASUREMENT_DEPTH in row {row} of sheet {sheet_name} is not an number"
                    f" as required")

            structure = data["required"]["structure"]
            validate_code(
                structure,
                "STRUCTURAL_FEATURE",
                "STRUCTURE",
                row,
                sheet_name,
                combined_concepts
            )

            dip = data["required"]["dip"]
            if not 0 >= dip >= -90:
                raise ConversionError(
                    f"The value {dip} for DIP in row {row} of sheet {sheet_name} is not between 0 and -90 as required")

            dip_direction = data["required"]["dip_direction"]
            if not -360 < dip_direction < 360:
                raise ConversionError(
                    f"The value {dip_direction} for DIP_DIRECTION in row {row} of sheet {sheet_name} is not between "
                    f"0 and 360 as required")

            alpha_angle = data["required"]["alpha_angle"]
            if not 0 < alpha_angle < 360:
                raise ConversionError(
                    f"The value {alpha_angle} for ALPHA_ANGLE in row {row} of sheet {sheet_name} is not between "
                    f"0 and 360 as required")
            
            beta_angle = data["required"]["beta_angle"]
            if not 0 < beta_angle < 360:
                raise ConversionError(
                    f"The value {beta_angle} for BETA_ANGLE in row {row} of sheet {sheet_name} is not between "
                    f"0 and 360 as required")

            azimuth = data["required"]["azimuth"]
            if not 0 <= azimuth <= 360:
                raise ConversionError(
                    f"The value {azimuth} for AZIMUTH in row {row} of sheet {sheet_name} is not between "
                    f"0 and 360 as required")

            if data["optional"].get("remark") is not None:
                remark = data["optional"]["remark"]

            # make RDFLib objects of the values
            drillhole_iri = URIRef(QLDBORES + drillhole_id)
            measurement_depth_lit = make_rdflib_type(measurement_depth, "Number")
            structure_iri = make_rdflib_type(structure, "Concept", combined_concepts)
            dip_lit = make_rdflib_type(dip, "Number")
            dip_direction_lit = make_rdflib_type(dip_direction, "Number")
            alpha_angle_lit = make_rdflib_type(alpha_angle, "Number")
            beta_angle_lit = make_rdflib_type(beta_angle, "Number")
            azimuth_lit = make_rdflib_type(azimuth, "Number")
            remark_lit = make_rdflib_type(remark, "String")
                
            # make the graph
            g.add((dataset_iri, SDO.hasPart, drillhole_iri))
            oc = BNode()
            g.add((oc, RDF.type, SOSA.ObservationCollection))
            g.add((oc, SOSA.hasFeatureOfInterest, drillhole_iri))
            g.add((drillhole_iri, SOSA.isFeatureOfInterestOf, oc))

            material_observations = [
                # name, op, value, unit, desc
                (Literal("Depth"), SDO.depth, measurement_depth_lit, UNITS.M, None),
                (Literal("Structure"), GEOSAMPLE.structure, structure_iri, None, None),
                (Literal("Dip"), BORE.hasDip, dip_lit, UNITS.DEG, None),
                (Literal("Dip Direction"), BORE.hasDipDirection, dip_direction_lit, UNITS.DEG, None),
                (Literal("Alpha Angle"), BORE.hasAlphaAngle, alpha_angle_lit, UNITS.DEG, None),
                (Literal("Beta Angle"), BORE.hasBetaAngle, beta_angle_lit, UNITS.DEG, None),
                (Literal("Azimuth"), BORE.hasAzimuth, azimuth_lit, UNITS.DEG, None),
            ]

            for n, op, v, u, d in material_observations:
                g += make_observation(oc, n, op, v, u, d)

            if remark_lit is not None:
                g.add((drillhole_iri, RDFS.comment, remark_lit))

            row += 1
        else:
            break

    g.bind("bore", BORE)
    g.bind("ex", EX)

    return g


def extract_sheet_surface_lithology(wb: openpyxl.Workbook, combined_concepts: Graph) -> Graph:
    check_template_version_supported(wb)

    sheet_name = "SURFACE_LITHOLOGY"
    sheet = wb[sheet_name]


def extract_sheet_surface_structure(wb: openpyxl.Workbook, combined_concepts: Graph) -> Graph:
    check_template_version_supported(wb)

    sheet_name = "SURFACE_STRUCTURE"
    sheet = wb[sheet_name]


def extract_sheet_reserves_resources(wb: openpyxl.Workbook, combined_concepts: Graph) -> Graph:
    check_template_version_supported(wb)

    sheet_name = "RESERVES_RESOURCES"
    sheet = wb[sheet_name]


def excel_to_rdf(
    file_to_convert_path: Path | BinaryIO,
    output_file_path: Optional[Path] = None
):
    """Converts a sheet within an Excel workbook to an RDF file"""
    wb = load_workbook(file_to_convert_path)
    template_version = get_template_version(wb)

    cc = Graph().parse(GSQ_PROFILE_DIR / "vocabs" / f"concepts-combined-{template_version}.ttl")

    # test that we have a valid template variable
    if template_version not in KNOWN_TEMPLATE_VERSIONS:
        raise ConversionError(
            f"Unknown Template Version. Known Template Versions are {', '.join(KNOWN_TEMPLATE_VERSIONS)},"
            f" you supplied {template_version}"
        )

    grf, dataset_iri = extract_sheet_dataset_metadata(wb, cc)
    grf: Graph
    dataset_iri: URIRef

    validate_sheet_validation_dictionary(wb, cc)
    grf += extract_sheet_user_dictionary(wb, template_version)
    validate_sheet_uom(wb, cc)
    g, uuo_notations = extract_sheet_user_uom(wb, cc)
    grf += g
    grf += extract_sheet_tenement(wb, cc, dataset_iri, template_version)
    g, drillhole_ids = extract_sheet_drillhole_location(wb, cc, dataset_iri, template_version)
    grf += g
    grf += extract_sheet_drillhole_survey(wb, cc, drillhole_ids, dataset_iri, template_version)
    g, sample_ids = extract_sheet_drillhole_sample(wb, cc, drillhole_ids, dataset_iri, template_version)
    grf += g
    g, sample_ids2 = extract_sheet_surface_sample(wb, cc, dataset_iri, template_version)
    grf += g
    sample_ids: []
    sample_ids += sample_ids2
    g, laboratories_dict = extract_sheet_user_laboratories(wb, dataset_iri, template_version)
    grf += g
    g, uspcs = extract_sheet_user_sample_prep_codes(wb, dataset_iri, template_version)
    grf += g
    g, assay_codes = extract_sheet_user_assay_codes(wb, dataset_iri, template_version)
    grf += g
    g, ans = extract_sheet_user_analytes(wb, dataset_iri, template_version)
    grf += g
    g, job_numbers = extract_sheet_sample_preparation(wb, laboratories_dict, uspcs, assay_codes, sample_ids, dataset_iri, template_version)
    grf += g
    uoms_concentration_notations = []
    for mem in cc.objects(URIRef("https://linked.data.gov.au/def/gsq-geochem/uom/concentration"), SKOS.member):
        uoms_concentration_notations.append(str(cc.value(subject=mem, predicate=SKOS.notation)))
    grf += extract_sheet_geochemistry_meta(wb, job_numbers, laboratories_dict, assay_codes, ans, uoms_concentration_notations, cc, dataset_iri, template_version)
    grf += extract_sheet_sample_geochemistry(wb, job_numbers, sample_ids, assay_codes, ans, dataset_iri, template_version)

    grf += extract_sheet_qaqc_meta(wb, job_numbers, laboratories_dict, assay_codes, ans, uoms_concentration_notations, cc, dataset_iri, template_version)
    grf += extract_sheet_qaqc_geochemistry(wb, job_numbers, sample_ids, assay_codes, ans, cc, dataset_iri, template_version)

    grf += extract_sheet_sample_pxrf(wb, sample_ids, ans, uoms_concentration_notations, cc, dataset_iri, template_version)

    g, lith_ids = extract_sheet_lith_dictionary(wb, URIRef("http://test.com"), "3.0")
    grf += g
    g, min_ids = extract_sheet_min_dictionary(wb, URIRef("http://test.com"), "3.0")
    grf += g
    grf += extract_sheet_drillhole_lithology(wb, drillhole_ids, lith_ids, min_ids, cc, dataset_iri, template_version)

    grf += extract_sheet_drillhole_structure(wb, drillhole_ids, cc, dataset_iri, template_version)

    grf.bind("bore", BORE)
    grf.bind("ex", EX)
    grf.bind(TENEMENT.prefix, TENEMENT)

    if output_file_path is not None:
        grf.serialize(destination=str(output_file_path), format="longturtle")
    else:  # print to std out
        return grf.serialize(format="longturtle")


def make_parser(args):
    parser = argparse.ArgumentParser(
        prog="geoexcelrdf", formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )

    parser.add_argument(
        "-i",
        "--info",
        help="The version and other info of this instance of Geochem Excel Converter",
        action="store_true",
    )

    parser.add_argument(
        "file_to_convert",
        nargs="?",  # allow 0 or 1 file name as argument
        type=Path,
        help="The Excel file to convert to a SKOS vocabulary in RDF or an RDF file to convert to an Excel file",
    )

    parser.add_argument(
        "-o",
        "--outputfile",
        help="An optionally-provided output file path. If not provided, output is to standard out",
        required=False,
    )

    parser.add_argument(
        "-e",
        "--external_metadata",
        help="Metadata for the Dataset object that is supplied outside the Excel file in an RDF file",
        type=Path,
        required=False,
        default=None
    )

    parser.add_argument(
        "-u",
        "--update_workbook",
        help="Update a given Excel Workbook's vocabularies",
        action="store_true",
    )

    return parser.parse_args(args)


def main(args=None):

    if args is None:  # run via entrypoint
        args = sys.argv[1:]

    args = make_parser(args)

    if not args:
        # show help if no args are given
        args.print_help()
        args.exit()

    elif args.info:
        from .__init__ import __version__

        print(f"geochemxl version: {__version__}")
        from .utils import KNOWN_TEMPLATE_VERSIONS

        print(
            f"Known template versions: {', '.join(sorted(KNOWN_TEMPLATE_VERSIONS, reverse=True))}"
        )
    elif args.update_workbook:
        print("Updating template")
        if args.file_to_convert is None:
            raise ValueError("If you select the option '-u', you must specify an Excel file to update")
        elif not Path(args.file_to_convert).is_file():
            raise ValueError("Files to update must exist")
        elif not args.file_to_convert.suffix.lower().endswith(tuple(EXCEL_FILE_ENDINGS)):
            raise ValueError("Files to update must end in .xslx")

        wb = load_workbook(args.file_to_convert)
        template_version = get_template_version(wb)

        # test that we have a valid template variable
        from .utils import KNOWN_TEMPLATE_VERSIONS
        if template_version not in KNOWN_TEMPLATE_VERSIONS:
            raise ValueError(
                f"Unknown Template Version. Known Template Versions are {', '.join(KNOWN_TEMPLATE_VERSIONS)},"
                f" you supplied {template_version}"
            )
        ws = wb["VALIDATION_DICTIONARY"]

        for vocab_id, vocab_path in FIELD_VOCABS.items():
            col = VOCAB_COLUMNS[vocab_id]
            row = 5
            for i in make_vocab_a_list_of_notations(vocab_path):
                ws[f"{col}{row}"] = i
                row += 1

        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1="=VALIDATION_DICTIONARY!$J$5:$J$13",
            showDropDown=False
        )
        ws2 = wb["DRILLHOLE_SAMPLE"]
        ws2.add_data_validation(dv)

        print(create_vocab_validation_formula(wb, "VALIDATION_DICTIONARY", "DRILL_TYPE"))
        print(create_vocab_validation_formula(wb, "DICTIONARY", "CODE"))

        wb.save(Path(args.file_to_convert).with_suffix(".y.xlsx"))
    elif args.file_to_convert:
        print(f"Processing file {args.file_to_convert}")

        # input file looks like an Excel file, so convert Excel -> RDF
        if not args.file_to_convert.suffix.lower().endswith(tuple(EXCEL_FILE_ENDINGS)):
            raise ConversionError("Only Excel files can be converted")
        else:
            try:
                o = excel_to_rdf(
                    args.file_to_convert,
                    output_file_path=args.outputfile,
                    external_metadata=args.external_metadata
                )
                if args.outputfile is None:
                    print(o)
            except ConversionError as err:
                logging.error("{0}".format(err))
                return 1
